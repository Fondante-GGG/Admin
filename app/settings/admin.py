import csv

from django import forms
from django.contrib import admin, messages
from django.contrib.auth import password_validation
from django.contrib.auth.admin import UserAdmin as DjangoUserAdmin
from django.contrib.auth.forms import ReadOnlyPasswordHashField
from django.contrib.admin.models import LogEntry
from django.contrib.contenttypes.models import ContentType
import io
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect
from django.http import HttpResponse, HttpResponseForbidden
from django.urls import path, reverse
from django.db import models
from django.db.models import Count, F, OuterRef, Prefetch, Q, Subquery, Sum, Value
from django.db.models.functions import Coalesce
from django.utils.dateparse import parse_date
from django.utils.html import format_html
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from django.utils.safestring import mark_safe
from django.utils import timezone
from datetime import date, timedelta
from django.http import JsonResponse
from django.template.response import TemplateResponse

from .admin_site import crm_admin_site
from .admin_site import _simple_pdf_table
from .receipts import payment_receipt_pdf
from .models import (
    AccountingAccount,
    AccountingCategory,
    AccountingEntry,
    AccountingProject,
    CalendarEvent,
    Call,
    CourseContract,
    Cursues,
    CourseDrop,
    CurriculumModule,
    DebtorEnrollment,
    Enrollment,
    GroupCourse,
    Lead,
    IndividualCourse,
    Lesson,
    Mentor,
    Organization,
    Parent,
    Payment,
    TuitionPayment,
    Salary,
    Student,
    StudentPayments,
    Task,
    User,
)

from django.core.exceptions import ValidationError


COURSE_LEGACY_STATUS_MAP = {
    "Подготовка курсов": "Подготовка",
    "Активные курсы": "Запущен",
    "Завершенные курсы": "Закончен",
}


def _course_status_choices():
    return list(Cursues._meta.get_field("status").choices)


def _course_status_values() -> list[str]:
    return [value for value, _label in _course_status_choices()]


def _canonical_course_status(status: str) -> str:
    return COURSE_LEGACY_STATUS_MAP.get(status, status)


def _course_status_filter_values(status: str) -> list[str]:
    values = [status]
    values.extend(legacy for legacy, current in COURSE_LEGACY_STATUS_MAP.items() if current == status)
    return values


def _role(user) -> str:
    return getattr(user, "role", "") or ""


LEAD_BOARD_COLUMNS = (
    {
        "key": "consultation",
        "title": "Консультация",
        "values": ("new",),
        "badge": "Важно",
        "tone": "warning",
    },
    {
        "key": "waiting",
        "title": "Ожидание",
        "values": ("in_progress",),
        "badge": "Важно",
        "tone": "warning",
    },
    {
        "key": "invited",
        "title": "Приглашен(а) на отк...",
        "values": ("invited",),
        "badge": "Важно",
        "tone": "warning",
    },
    {
        "key": "participated",
        "title": "Участвовал(а) на от...",
        "values": ("participated",),
        "badge": "Успешно",
        "tone": "success",
    },
    {
        "key": "studying",
        "title": "УЖЕ УЧИТЬСЯ",
        "values": ("won",),
        "badge": "Успешно",
        "tone": "success",
    },
    {
        "key": "lost",
        "title": "Потерянный лид",
        "values": ("lost",),
        "badge": "Потерян",
        "tone": "muted",
    },
)

LEAD_STATUS_CHOICES = (
    ("new", "Консультация"),
    ("in_progress", "Ожидание"),
    ("invited", "Приглашен(а) на открытый урок"),
    ("participated", "Участвовал(а) на открытом уроке"),
    ("won", "Уже учится"),
    ("lost", "Потерянный лид"),
)

LEAD_STATUS_LABELS = dict(LEAD_STATUS_CHOICES)
LEAD_STATUS_TO_COLUMN = {
    value: column["key"]
    for column in LEAD_BOARD_COLUMNS
    for value in column["values"]
}

LEAD_EXPORT_FIELDS = (
    ("id", "ID", True),
    ("first_name", "Имя", True),
    ("last_name", "Фамилия", True),
    ("middle_name", "Отчество", True),
    ("phone_number", "Номер телефона", True),
    ("extra_phone", "Дополнительный номер", False),
    ("email", "Адрес электронной почты", True),
    ("created_at", "Дата добавления", False),
    ("channel", "Канал", True),
    ("status", "Статус", True),
    ("assignee", "Назначен на", False),
    ("archived_at", "Дата архивирования", False),
    ("lost_reason", "Причина потери лида", False),
    ("birth_date", "Дата рождения", False),
    ("due_date", "Крайний срок", False),
    ("subject_interest", "Интересуется предметом", False),
    ("source", "С какой публичной формы", False),
    ("interested_courses", "Интересуется курсами", False),
    ("telegram_nick", "Ник в Telegram", False),
    ("from_where", "Откуда", False),
    ("labels", "Метки", False),
    ("comment", "Комментарий", False),
)


class LeadAdminForm(forms.ModelForm):
    status = forms.ChoiceField(choices=LEAD_STATUS_CHOICES, label="Статус")

    class Meta:
        model = Lead
        fields = "__all__"


class CRMUserCreationForm(forms.ModelForm):
    password1 = forms.CharField(
        label="Пароль",
        strip=False,
        widget=forms.PasswordInput,
        help_text=password_validation.password_validators_help_text_html(),
    )
    password2 = forms.CharField(
        label="Подтверждение пароля",
        strip=False,
        widget=forms.PasswordInput,
    )

    class Meta:
        model = User
        fields = (
            "username",
            "first_name",
            "last_name",
            "email",
            "phone_number",
            "role",
            "is_staff",
            "is_active",
            "is_superuser",
        )

    def clean_password2(self):
        password1 = self.cleaned_data.get("password1")
        password2 = self.cleaned_data.get("password2")
        if password1 and password2 and password1 != password2:
            raise ValidationError("Пароли не совпадают.")
        probe_user = User(
            username=self.cleaned_data.get("username", ""),
            first_name=self.cleaned_data.get("first_name", ""),
            last_name=self.cleaned_data.get("last_name", ""),
            email=self.cleaned_data.get("email", ""),
            phone_number=self.cleaned_data.get("phone_number", ""),
            role=self.cleaned_data.get("role", ""),
            is_staff=self.cleaned_data.get("is_staff", False),
            is_active=self.cleaned_data.get("is_active", True),
            is_superuser=self.cleaned_data.get("is_superuser", False),
        )
        password_validation.validate_password(password2, probe_user)
        return password2

    def save(self, commit=True):
        user = super().save(commit=False)
        user.set_password(self.cleaned_data["password1"])
        if commit:
            user.save()
            self.save_m2m()
        return user


class CRMUserChangeForm(forms.ModelForm):
    password = ReadOnlyPasswordHashField(
        label="Текущий пароль",
        help_text="Сырый пароль не хранится. Чтобы сменить пароль, задайте новый ниже.",
    )
    new_password1 = forms.CharField(
        label="Новый пароль",
        required=False,
        strip=False,
        widget=forms.PasswordInput,
        help_text=password_validation.password_validators_help_text_html(),
    )
    new_password2 = forms.CharField(
        label="Подтверждение нового пароля",
        required=False,
        strip=False,
        widget=forms.PasswordInput,
    )

    class Meta:
        model = User
        fields = "__all__"

    def clean_password(self):
        return self.initial.get("password")

    def clean(self):
        cleaned_data = super().clean()
        new_password1 = cleaned_data.get("new_password1")
        new_password2 = cleaned_data.get("new_password2")
        if new_password1 or new_password2:
            if new_password1 != new_password2:
                raise ValidationError("Новые пароли не совпадают.")
            password_validation.validate_password(new_password2, self.instance)
        return cleaned_data

    def save(self, commit=True):
        user = super().save(commit=False)
        new_password = self.cleaned_data.get("new_password1")
        if new_password:
            user.set_password(new_password)
        if commit:
            user.save()
            self.save_m2m()
        return user


class RoleRestrictedAdminMixin:
    allowed_roles: set[str] | None = None
    denied_roles: set[str] | None = None

    def _is_allowed(self, request) -> bool:
        role = _role(request.user)
        if role == "Админ":
            role = "Администратор"
        if self.denied_roles and role in self.denied_roles:
            return False
        if self.allowed_roles is None:
            return True
        return role in self.allowed_roles

    def _has_crm_access(self, request) -> bool:
        user = request.user
        return bool(user.is_authenticated and user.is_active and user.is_staff and self._is_allowed(request))

    def has_module_permission(self, request):
        return self._has_crm_access(request)

    def has_view_permission(self, request, obj=None):
        return self._has_crm_access(request)

    def has_add_permission(self, request):
        return self._has_crm_access(request)

    def has_change_permission(self, request, obj=None):
        return self._has_crm_access(request)

    def has_delete_permission(self, request, obj=None):
        return self._has_crm_access(request)

    def get_model_perms(self, request):
        if not self.has_module_permission(request):
            return {}
        return {
            "add": self.has_add_permission(request),
            "change": self.has_change_permission(request),
            "delete": self.has_delete_permission(request),
            "view": self.has_view_permission(request),
        }


class ArchiveFilter(admin.SimpleListFilter):
    title = "Архив"
    parameter_name = "archived"

    def lookups(self, request, model_admin):
        return (("0", "Активные"), ("1", "Архив"))

    def queryset(self, request, queryset):
        v = self.value()
        if v == "1":
            return queryset.filter(is_archived=True)
        if v == "0":
            return queryset.filter(is_archived=False)
        return queryset


def archive_selected(modeladmin, request, queryset):
    queryset.update(is_archived=True, archived_at=timezone.now())


archive_selected.short_description = "Переместить в архив"


def unarchive_selected(modeladmin, request, queryset):
    queryset.update(is_archived=False, archived_at=None)


unarchive_selected.short_description = "Вернуть из архива"


class ArchiveAdminMixin:
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        # default: show active only
        if request.GET.get("archived") == "1":
            return qs
        if request.GET.get("archived") == "0" or "archived" not in request.GET:
            return qs.filter(is_archived=False)
        return qs


class OrganizationFilterMixin:
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        org_id = request.session.get("current_org_id")
        if org_id and hasattr(self.model, "organization"):
            qs = qs.filter(organization_id=org_id)
        return qs


class OrganizationFilter(admin.SimpleListFilter):
    title = "Организация"
    parameter_name = "organization"

    def lookups(self, request, model_admin):
        return ((str(o.pk), o.name) for o in Organization.objects.all().order_by("name"))

    def queryset(self, request, queryset):
        v = self.value()
        if v:
            return queryset.filter(organization_id=v)
        return queryset


@admin.register(Organization, site=crm_admin_site)
class OrganizationAdmin(RoleRestrictedAdminMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор"}
    list_display = ("name", "slug", "created_at")
    search_fields = ("name", "slug")
    prepopulated_fields = {"slug": ("name",)}


@admin.register(Parent, site=crm_admin_site)
class ParentAdmin(RoleRestrictedAdminMixin, OrganizationFilterMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор", "Менеджер"}
    change_list_template = "admin/parents_changelist.html"
    list_display = ("user", "phone_number", "students_list", "created_at")
    list_filter = ("is_archived", OrganizationFilter)
    search_fields = ("user__username", "user__first_name", "user__last_name", "phone_number")
    autocomplete_fields = ("user", "students")
    filter_horizontal = ("students",)

    def students_list(self, obj: Parent):
        return ", ".join(str(s.user.get_full_name() or s.user.username) for s in obj.students.all()[:5]) or "—"
    students_list.short_description = "Студенты"

    @staticmethod
    def _students_for_parent_form(request):
        qs = Student.objects.select_related("user").filter(is_archived=False)
        org_id = request.session.get("current_org_id")
        if org_id:
            qs = qs.filter(organization_id=org_id)
        return qs.order_by("user__last_name", "user__first_name", "user__username")

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context["students_for_parent"] = self._students_for_parent_form(request)
        return super().changelist_view(request, extra_context=extra_context)


@admin.register(User, site=crm_admin_site)
class UserAdmin(RoleRestrictedAdminMixin, DjangoUserAdmin):
    allowed_roles = {"Администратор", "Менеджер"}
    form = CRMUserChangeForm
    add_form = CRMUserCreationForm
    list_display = ("username", "first_name", "last_name", "email", "phone_number", "role", "is_staff", "is_active")
    list_filter = ("role", "is_staff", "is_active")
    search_fields = ("username", "first_name", "last_name", "email", "phone_number")
    ordering = ("username",)
    fieldsets = (
        (None, {"fields": ("username", "password")}),
        ("Личные данные", {"fields": ("first_name", "last_name", "email", "phone_number", "role")}),
        ("Смена пароля", {"fields": ("new_password1", "new_password2")}),
        ("Права доступа", {"fields": ("is_active", "is_staff", "is_superuser", "groups", "user_permissions")}),
        ("Важные даты", {"fields": ("last_login", "date_joined")}),
    )
    add_fieldsets = (
        (
            None,
            {
                "classes": ("wide",),
                "fields": (
                    "username",
                    "first_name",
                    "last_name",
                    "email",
                    "phone_number",
                    "role",
                    "password1",
                    "password2",
                    "is_staff",
                    "is_active",
                    "is_superuser",
                ),
            },
        ),
    )


@admin.register(Student, site=crm_admin_site)
class StudentAdmin(RoleRestrictedAdminMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор", "Менеджер", "Ментор"}
    change_list_template = "admin/students_changelist.html"
    list_display = (
        "student_id",
        "full_name",
        "group_display",
        "paid_total_display",
        "status_badge",
        "account_display",
    )
    list_display_links = None
    search_fields = ("user__username", "user__first_name", "user__last_name", "user__phone_number")
    autocomplete_fields = ("user",)
    list_per_page = 20
    list_filter = ("status", ArchiveFilter)
    actions = (archive_selected, unarchive_selected)
    change_form_template = "admin/student_change_form.html"

    @staticmethod
    def _safe_next_url(request):
        next_url = (request.POST.get("next") or request.GET.get("next") or "").strip()
        if next_url.startswith("/") and not next_url.startswith("//"):
            return next_url
        return ""

    @staticmethod
    def _course_id_from_request(request):
        course_id = (request.POST.get("course") or request.GET.get("course") or "").strip()
        return int(course_id) if course_id.isdigit() else None

    def _course_from_request(self, request):
        course_id = self._course_id_from_request(request)
        if course_id is None:
            return None
        return Cursues.objects.filter(pk=course_id, is_archived=False).first()

    def get_queryset(self, request):
        qs = super().get_queryset(request).select_related("user")
        pay_sub = (
            Payment.objects.filter(student_id=OuterRef("pk"))
            .values("student_id")
            .annotate(total=Sum("amount"))
            .values("total")[:1]
        )
        qs = qs.annotate(
            paid_total=Coalesce(
                Subquery(pay_sub),
                Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2)),
                output_field=models.DecimalField(max_digits=12, decimal_places=2),
            )
        )

        pay_course = request.GET.get("pay_course")
        if pay_course:
            qs = qs.filter(payment__course_id=pay_course)

        pay_mentor = request.GET.get("pay_mentor")
        if pay_mentor:
            qs = qs.filter(payment__course__mentors__id=pay_mentor)

        pay_method = request.GET.get("pay_method")
        if pay_method:
            qs = qs.filter(payment__method=pay_method)

        pay_id = (request.GET.get("pay_id") or "").strip()
        if pay_id:
            qs = qs.filter(payment__id=pay_id)

        pay_from = parse_date(request.GET.get("pay_date_from") or "")
        if pay_from:
            qs = qs.filter(payment__created_at__date__gte=pay_from)

        pay_to = parse_date(request.GET.get("pay_date_to") or "")
        if pay_to:
            qs = qs.filter(payment__created_at__date__lte=pay_to)

        pay_first = request.GET.get("pay_first")
        if pay_first == "1":
            qs = qs.filter(payment__isnull=False)
        elif pay_first == "0":
            qs = qs.filter(payment__isnull=True)

        qs = qs.distinct()

        if request.GET.get("archived") == "1":
            return qs.filter(is_archived=True)
        if request.GET.get("archived") == "0" or "archived" not in request.GET:
            return qs.filter(is_archived=False)
        return qs

    def student_id(self, obj: Student):
        return format_html(
            '<a href="#" class="crm-student-link" data-student-id="{}">#{}</a>',
            obj.pk,
            obj.pk,
        )

    student_id.short_description = "ID"

    def full_name(self, obj: Student):
        name = obj.user.get_full_name() or obj.user.username
        return format_html(
            '<a href="#" class="crm-student-link" data-student-id="{}">{}</a>',
            obj.pk,
            name,
        )

    full_name.short_description = "ФИО"

    def group_display(self, obj: Student):
        # show any related course title (latest enrollment), fallback empty
        enroll = (
            Enrollment.objects.select_related("course")
            .filter(student=obj, course__is_archived=False)
            .order_by("-created_at")
            .first()
        )
        return enroll.course.title if enroll else "—"

    group_display.short_description = "Группа"

    def status_badge(self, obj: Student):
        color = {
            Student.Status.ACTIVE: "#10b981",
            Student.Status.INACTIVE: "#6b7280",
            Student.Status.LEFT: "#f59e0b",
            Student.Status.FROZEN: "#a855f7",
        }.get(obj.status, "#6b7280")
        label = obj.get_status_display()
        return format_html(
            '<span style="display:inline-block;padding:2px 10px;border-radius:999px;background:{}20;color:{};font-weight:800;">{}</span>',
            color,
            color,
            label,
        )

    status_badge.short_description = "Статус"

    def paid_total_display(self, obj: Student):
        return f"{int(obj.paid_total or 0):,}".replace(",", " ")

    paid_total_display.short_description = "Оплачено (с.)"

    def account_display(self, obj: Student):
        return "Да" if obj.user.is_active else "Нет"

    account_display.short_description = "Аккаунт"

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        qs = self.get_queryset(request)
        extra_context["courses"] = Cursues.objects.filter(is_archived=False).order_by("title")
        extra_context["mentors"] = Mentor.objects.select_related("user").order_by("user__username")
        extra_context["pay_methods"] = Payment.Method.choices
        extra_context["pay_filters"] = {
            "pay_course": request.GET.get("pay_course", ""),
            "pay_mentor": request.GET.get("pay_mentor", ""),
            "pay_method": request.GET.get("pay_method", ""),
            "pay_first": request.GET.get("pay_first", ""),
            "pay_date_from": request.GET.get("pay_date_from", ""),
            "pay_date_to": request.GET.get("pay_date_to", ""),
            "pay_id": request.GET.get("pay_id", ""),
        }
        extra_context["student_counts"] = {
            "active": qs.filter(status=Student.Status.ACTIVE).count(),
            "inactive": qs.filter(status=Student.Status.INACTIVE).count(),
            "left": qs.filter(status=Student.Status.LEFT).count(),
            "frozen": qs.filter(status=Student.Status.FROZEN).count(),
        }
        extra_context["students_paid_total"] = float(
            Payment.objects.filter(student_id__in=qs.values("id")).aggregate(total=Sum("amount"))["total"] or 0
        )
        return super().changelist_view(request, extra_context=extra_context)

    def changeform_view(self, request, object_id=None, form_url="", extra_context=None):
        extra_context = extra_context or {}
        course = self._course_from_request(request)
        if course:
            extra_context["enroll_course"] = course
            extra_context["next_url"] = self._safe_next_url(request)
        return super().changeform_view(request, object_id, form_url, extra_context=extra_context)

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        course = self._course_from_request(request)
        if course:
            Enrollment.objects.get_or_create(
                student=obj,
                course=course,
                defaults={"tuition_amount": course.price or 0},
            )
            course.students.add(obj)

    def delete_queryset(self, request, queryset):
        student_ids = list(queryset.values_list("pk", flat=True))
        if student_ids:
            Student.objects.filter(pk__in=student_ids).delete()

    def response_add(self, request, obj, post_url_continue=None):
        next_url = self._safe_next_url(request)
        if next_url and "_continue" not in request.POST and "_addanother" not in request.POST:
            return redirect(next_url)
        return super().response_add(request, obj, post_url_continue=post_url_continue)


class StudentEnrollmentInline(admin.TabularInline):
    model = Enrollment
    extra = 0
    autocomplete_fields = ("course",)


class StudentPaymentInline(admin.TabularInline):
    model = Payment
    extra = 0
    autocomplete_fields = ("course",)


class StudentCourseDropInline(admin.TabularInline):
    model = CourseDrop
    extra = 0
    autocomplete_fields = ("course",)


StudentAdmin.inlines = [StudentEnrollmentInline, StudentPaymentInline, StudentCourseDropInline]


@admin.register(Mentor, site=crm_admin_site)
class MentorAdmin(RoleRestrictedAdminMixin, OrganizationFilterMixin, ArchiveAdminMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор", "Менеджер"}
    change_list_template = "admin/mentors_changelist.html"
    fieldsets = (
        (None, {"fields": ("user",)}),
        (
            "Личные данные",
            {
                "fields": (
                    "middle_name",
                    "birth_date",
                    "skills",
                    "workplace",
                    "documents_folder",
                )
            },
        ),
        (
            "Оплата",
            {"fields": ("payment_form", "payment_rate", "percentage_rate", "fixed_rate")},
        ),
        ("Контракт и примечания", {"fields": ("contract_file", "note")}),
        ("Уход", {"fields": ("departure_date", "departure_reason")}),
        ("Системное", {"fields": ("created_at",)}),
    )
    readonly_fields = ("created_at",)
    list_display = (
        "mentor_id_display",
        "full_name_display",
        "course_display",
        "status_badge_display",
        "account_display",
        "phone_display",
        "email_display",
    )
    list_display_links = None
    search_fields = (
        "user__username",
        "user__first_name",
        "user__last_name",
        "user__phone_number",
        "user__email",
    )
    list_filter = (ArchiveFilter,)
    actions = (archive_selected, unarchive_selected)
    autocomplete_fields = ("user",)
    list_per_page = 20

    def get_queryset(self, request):
        course_qs = Cursues.objects.filter(is_archived=False).only("title", "id")
        qs = (
            super()
            .get_queryset(request)
            .select_related("user")
            .prefetch_related(Prefetch("cursues_set", queryset=course_qs))
        )
        if request.GET.get("archived") == "1":
            return qs
        if request.GET.get("archived") == "0" or "archived" not in request.GET:
            return qs.filter(is_archived=False)
        return qs

    def get_search_results(self, request, queryset, search_term):
        if not search_term:
            return queryset, False
        q = search_term.strip()
        text_q = (
            Q(user__first_name__icontains=q)
            | Q(user__last_name__icontains=q)
            | Q(user__username__icontains=q)
            | Q(user__phone_number__icontains=q)
            | Q(user__email__icontains=q)
        )
        if q.isdigit():
            queryset = queryset.filter(Q(pk=int(q)) | text_q)
        else:
            queryset = queryset.filter(text_q)
        return queryset, False

    def mentor_id_display(self, obj: Mentor):
        url = reverse(f"{crm_admin_site.name}:settings_mentor_change", args=[obj.pk])
        return format_html('<a href="{}" class="crm-mentor-table-link" data-mentor-id="{}">#{}</a>', url, obj.pk, obj.pk)

    mentor_id_display.short_description = "ID"

    def full_name_display(self, obj: Mentor):
        u = obj.user
        parts = [u.last_name or "", u.first_name or "", obj.middle_name or ""]
        name = " ".join(p for p in parts if p).strip() or u.get_full_name() or u.username
        url = reverse(f"{crm_admin_site.name}:settings_mentor_change", args=[obj.pk])
        return format_html('<a href="{}" class="crm-mentor-table-link" data-mentor-id="{}">{}</a>', url, obj.pk, name)

    full_name_display.short_description = "ФИО"

    def course_display(self, obj: Mentor):
        titles = [c.title for c in obj.cursues_set.all()[:15]]
        return ", ".join(titles) if titles else "—"

    course_display.short_description = "Курс"

    def status_badge_display(self, obj: Mentor):
        u = obj.user
        if u.is_active:
            return format_html(
                '<span class="crm-mentor-status crm-mentor-status--active">Активный</span>'
            )
        return format_html(
            '<span class="crm-mentor-status crm-mentor-status--inactive">Неактивен</span>'
        )

    status_badge_display.short_description = "Статус"

    def account_display(self, obj: Mentor):
        return "Да" if obj.user.is_active else "Нет"

    account_display.short_description = "Аккаунт"

    def phone_display(self, obj: Mentor):
        return obj.user.phone_number or "—"

    phone_display.short_description = "Телефон"

    def email_display(self, obj: Mentor):
        return obj.user.email or "—"

    email_display.short_description = "Почта"

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context["mentor_total"] = Mentor.objects.count()
        extra_context["mentors_export_url"] = reverse(f"{crm_admin_site.name}:mentors_export_xlsx")
        extra_context["mentors_salary_url"] = reverse(f"{crm_admin_site.name}:mentors_salary")
        extra_context["mentor_create_url"] = reverse(f"{crm_admin_site.name}:mentor_quick_create")
        extra_context["mentor_payment_form_choices"] = Mentor.PaymentForm.choices
        return super().changelist_view(request, extra_context=extra_context)


@admin.register(Cursues, site=crm_admin_site)
class CursuesAdmin(RoleRestrictedAdminMixin, OrganizationFilterMixin, ArchiveAdminMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор", "Менеджер", "Ментор"}
    list_display = ("title", "course_type", "subject", "start", "end", "lessons_per_month", "status", "price", "students_badge")
    list_filter = ("course_type", "status", "subject", ArchiveFilter)
    search_fields = ("title",)
    filter_horizontal = ("students", "mentors")
    change_form_template = "admin/course_change_form.html"
    actions = (archive_selected, unarchive_selected)
    list_per_page = 20
    fieldsets = (
        (None, {"fields": ("title", "course_type", "subject", "status")}),
        ("Сроки", {"fields": ("start", "end", "duration_days", "lessons_per_month")}),
        ("Детали", {"fields": ("price", "capacity", "room", "schedule_note")}),
        ("Участники", {"fields": ("mentors", "students")}),
        ("Организация", {"fields": ("organization",)}),
    )

    def students_badge(self, obj: Cursues):
        count = obj.students.count()
        capacity = obj.capacity if obj.capacity is not None else "—"
        return mark_safe(
            f'<span style="display:inline-block;padding:2px 10px;border-radius:999px;background:#ef4444;color:#fff;font-weight:800;">{count}/{capacity}</span>'
        )

    students_badge.short_description = "Студенты"

    def save_model(self, request, obj, form, change):
        if obj.status == "Архивирован":
            obj.is_archived = True
            obj.archived_at = obj.archived_at or timezone.now()
        elif obj.is_archived and obj.status != "Архивирован" and "status" in getattr(form, "changed_data", []):
            obj.is_archived = False
            obj.archived_at = None
        super().save_model(request, obj, form, change)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        extra_context = extra_context or {}
        course = self.get_object(request, object_id)
        if course:
            enrollments = (
                Enrollment.objects
                .filter(course=course)
                .select_related("student__user")
                .order_by("student__user__first_name", "student__user__username")
            )
            contract_map = {
                contract.student_id: contract
                for contract in CourseContract.objects.filter(course=course).select_related("student")
            }
            for e in enrollments:
                e.paid_amount = e.paid_total
                e.debt_amount = e.debt
                e.contract = contract_map.get(e.student_id)

            extra_context["course_enrollments"] = enrollments
            extra_context["course_students_count"] = enrollments.count()
            extra_context["course_capacity"] = course.capacity

            payments = (
                Payment.objects
                .filter(course=course)
                .select_related("student__user")
                .order_by("-created_at")
            )
            extra_context["course_payments"] = payments
            extra_context["course_payments_total"] = float(
                payments.aggregate(total=Sum("amount"))["total"] or 0
            )
            extra_context["course_tuition_total"] = float(
                enrollments.aggregate(total=Sum("tuition_amount"))["total"] or 0
            )
            extra_context["course_debt_total"] = float(
                (extra_context["course_tuition_total"] or 0) - (extra_context["course_payments_total"] or 0)
            )
            extra_context["course_duration_label"] = getattr(course, "duration_label", "—")
            extra_context["course_drops"] = (
                CourseDrop.objects
                .filter(course=course)
                .select_related("student__user")
                .order_by("-dropped_at", "student__user__first_name", "student__user__username")
            )
            extra_context["course_status_value"] = _canonical_course_status(course.status) if course.status else ""
            extra_context["course_status_label"] = extra_context["course_status_value"] or "—"
            extra_context["course_duration_months"] = max(1, round((course.duration_days or 0) / 30)) if (course.duration_days or 0) > 0 else 1
            extra_context["course_update_url"] = reverse(f"{crm_admin_site.name}:course_update", args=[course.pk])
            extra_context["course_status_url"] = reverse(f"{crm_admin_site.name}:course_status_update", args=[course.pk])
            extra_context["course_return_url"] = reverse(f"{crm_admin_site.name}:settings_groupcourse_change", args=[course.pk])
            extra_context["course_students_csv_url"] = reverse(f"{crm_admin_site.name}:course_students_csv", args=[course.pk])
            extra_context["course_students_upload_url"] = reverse(f"{crm_admin_site.name}:course_students_upload", args=[course.pk])
            extra_context["course_contracts_generate_url"] = reverse(f"{crm_admin_site.name}:course_contracts_generate", args=[course.pk])
            extra_context["course_payment_methods"] = Payment.Method.choices
            extra_context["course_status_choices"] = _course_status_choices()
            extra_context["course_subject_choices"] = list(
                Cursues.objects.exclude(subject="").values_list("subject", flat=True).distinct().order_by("subject")
            )
            extra_context["course_student_add_url"] = reverse(f"{crm_admin_site.name}:settings_student_add")

        return super().change_view(request, object_id, form_url, extra_context)


@admin.register(CurriculumModule, site=crm_admin_site)
class CurriculumModuleAdmin(RoleRestrictedAdminMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор", "Менеджер"}
    list_display = ("course", "order", "title")
    search_fields = ("title", "course__title")
    list_filter = ("course",)
    ordering = ("course", "order")


class CurriculumModuleInline(admin.TabularInline):
    model = CurriculumModule
    extra = 0
    fields = ("order", "title")
    ordering = ("order",)

    def has_add_permission(self, request, obj=None):
        role = _role(request.user)
        if role == "Админ":
            role = "Администратор"
        return role in {"Администратор", "Менеджер"}

    def has_change_permission(self, request, obj=None):
        return self.has_add_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        return self.has_add_permission(request, obj)


class EnrollmentInline(admin.TabularInline):
    model = Enrollment
    extra = 0
    autocomplete_fields = ("student",)


class CourseDropInline(admin.TabularInline):
    model = CourseDrop
    extra = 0
    autocomplete_fields = ("student",)


CursuesAdmin.inlines = [CurriculumModuleInline, EnrollmentInline, CourseDropInline]


@admin.register(GroupCourse, site=crm_admin_site)
class GroupCourseAdmin(CursuesAdmin):
    allowed_roles = {"Администратор", "Менеджер", "Ментор"}
    change_list_template = "admin/group_courses_changelist.html"
    change_form_template = "admin/group_course_change_form.html"

    def save_model(self, request, obj, form, change):
        obj.course_type = Cursues.CourseType.GROUP
        super().save_model(request, obj, form, change)

    def get_queryset(self, request):
        qs = super().get_queryset(request).prefetch_related("students", "mentors__user")
        return qs.filter(course_type=Cursues.CourseType.GROUP)

    list_per_page = 20

    @staticmethod
    def _course_card(course: Cursues) -> dict:
        mentors = list(course.mentors.all())
        mentor = mentors[0] if mentors else None
        mentor_user = mentor.user if mentor else None
        students_count = len(course.students.all())

        end_date = None
        if course.start and course.duration_days:
            end_date = course.start + timedelta(days=max(int(course.duration_days) - 1, 0))

        badge_bg = "#adff2f" if course.capacity and students_count >= course.capacity else "#ff5757"
        badge_color = "#15212f" if course.capacity and students_count >= course.capacity else "#ffffff"

        return {
            "id": course.pk,
            "title": course.title,
            "room": (course.room or "—").strip() or "—",
            "mentor_name": (mentor_user.get_full_name() or mentor_user.username) if mentor_user else "—",
            "start": course.start,
            "end": end_date,
            "schedule_note": (course.schedule_note or "").strip(),
            "duration_label": getattr(course, "duration_label", "—"),
            "students_count": students_count,
            "capacity": course.capacity,
            "badge_bg": badge_bg,
            "badge_color": badge_color,
            "change_url": reverse(f"{crm_admin_site.name}:settings_groupcourse_change", args=[course.pk]),
        }

    def changelist_view(self, request, extra_context=None):
        request._crm_group_params = request.GET.copy()
        mutable = request.GET.copy()
        for key in ("view", "group_by"):
            mutable.pop(key, None)
        request.GET = mutable
        request.META["QUERY_STRING"] = mutable.urlencode()

        extra_context = extra_context or {}
        base = self.get_queryset(request)
        params = request._crm_group_params

        q = (params.get("q") or "").strip()
        if q:
            base = base.filter(title__icontains=q)

        subject = params.get("subject", "")
        if subject:
            base = base.filter(subject=subject)

        status = params.get("status", "")
        if status:
            base = base.filter(status__in=_course_status_filter_values(status))

        group_by = params.get("group_by", "status")
        statuses = _course_status_values()

        def build_url(**updates):
            params = request._crm_group_params.copy()
            for key, value in updates.items():
                if value in (None, "", False):
                    params.pop(key, None)
                else:
                    params[key] = str(value)
            qs = params.urlencode()
            return f"{request.path}?{qs}" if qs else request.path

        extra_context["view_mode"] = params.get("view", "grid")
        extra_context["group_by"] = group_by
        extra_context["selected_status"] = params.get("status", "")
        extra_context["selected_subject"] = params.get("subject", "")
        extra_context["search_query"] = q
        extra_context["group_current_url"] = build_url()

        subjects = list(
            self.get_queryset(request)
            .exclude(subject="")
            .values_list("subject", flat=True)
            .distinct()
            .order_by("subject")
        )
        extra_context["subjects"] = subjects
        extra_context["status_choices"] = statuses
        extra_context["course_status_choices"] = _course_status_choices()
        extra_context["course_subject_choices"] = subjects
        extra_context["course_create_url"] = reverse(f"{crm_admin_site.name}:course_quick_create")

        if group_by == "subject":
            sections = []
            for subj in subjects:
                items = list(base.filter(subject=subj).order_by("-created_at"))
                if not items:
                    continue
                sections.append(
                    {
                        "title": subj,
                        "key": subj,
                        "items": [self._course_card(course) for course in items],
                    }
                )
            if not sections:
                no_subject_items = list(base.filter(subject="").order_by("-created_at"))
                if no_subject_items:
                    sections.append(
                        {
                            "title": "Без предмета",
                            "key": "without-subject",
                            "items": [self._course_card(course) for course in no_subject_items],
                        }
                    )
        else:
            sections = []
            for s in statuses:
                items = list(base.filter(status__in=_course_status_filter_values(s)).order_by("-created_at"))
                if not items:
                    continue
                sections.append(
                    {
                        "title": s,
                        "key": s,
                        "items": [self._course_card(course) for course in items],
                    }
                )

        extra_context["status_sections"] = sections
        extra_context["group_add_url"] = reverse(
            f"{self.admin_site.name}:{self.opts.app_label}_{self.opts.model_name}_add"
        )
        is_archive = params.get("archived") == "1"
        extra_context["group_archive_url"] = build_url(archived="" if is_archive else "1")
        extra_context["group_is_archive"] = is_archive
        extra_context["group_grid_url"] = build_url(view="grid")
        extra_context["group_list_url"] = build_url(view="list")
        extra_context["group_status_url"] = build_url(group_by="status")
        extra_context["group_subject_url"] = build_url(group_by="subject")
        return super().changelist_view(request, extra_context=extra_context)


@admin.register(IndividualCourse, site=crm_admin_site)
class IndividualCourseAdmin(CursuesAdmin):
    allowed_roles = {"Администратор", "Менеджер", "Ментор"}
    change_list_template = "admin/individual_courses_changelist.html"

    def save_model(self, request, obj, form, change):
        obj.course_type = Cursues.CourseType.INDIVIDUAL
        super().save_model(request, obj, form, change)

    def get_queryset(self, request):
        qs = super().get_queryset(request).prefetch_related("students__user", "mentors__user")
        return qs.filter(course_type=Cursues.CourseType.INDIVIDUAL)

    list_display = ("id", "title", "student_display", "mentor_display", "start", "status")
    list_display_links = ("title",)
    ordering = ("-created_at",)
    list_per_page = 20

    @staticmethod
    def _status_meta(status: str) -> tuple[str, str]:
        status = _canonical_course_status(status)
        mapping = {
            "Подготовка": ("#f59e0b", "#fff7ed"),
            "Готов к запуску": ("#2563eb", "#eff6ff"),
            "Запущен": ("#10b981", "#ecfdf5"),
            "Приостановлен": ("#ef4444", "#fef2f2"),
            "Закончен": ("#6b7280", "#f3f4f6"),
            "Архивирован": ("#64748b", "#f1f5f9"),
        }
        color, bg = mapping.get(status, ("#6c5ce7", "#f3f0ff"))
        return color, bg

    def student_display(self, obj: Cursues):
        st = obj.students.select_related("user").first()
        if not st:
            return "—"
        user = st.user
        return user.get_full_name() or user.username

    student_display.short_description = "Студент"

    def mentor_display(self, obj: Cursues):
        m = obj.mentors.select_related("user").first()
        if not m:
            return "—"
        user = m.user
        return user.get_full_name() or user.username

    mentor_display.short_description = "Ментор"

    def changelist_view(self, request, extra_context=None):
        response = super().changelist_view(request, extra_context=extra_context)
        if not hasattr(response, "context_data") or "cl" not in response.context_data:
            return response

        cl = response.context_data["cl"]
        row_items = []
        for course in cl.result_list:
            student = course.students.all().first()
            student_user = student.user if student else None
            mentor = course.mentors.all().first()
            mentor_user = mentor.user if mentor else None
            status_color, status_bg = self._status_meta(course.status)
            row_items.append(
                {
                    "id": course.pk,
                    "title": course.title,
                    "student": (student_user.get_full_name() or student_user.username) if student_user else "—",
                    "mentor": (mentor_user.get_full_name() or mentor_user.username) if mentor_user else "—",
                    "start": course.start,
                    "status": _canonical_course_status(course.status),
                    "status_color": status_color,
                    "status_bg": status_bg,
                    "change_url": reverse(
                        f"{self.admin_site.name}:{self.opts.app_label}_{self.opts.model_name}_change",
                        args=[course.pk],
                    ),
                }
            )

        add_url = reverse(f"{self.admin_site.name}:{self.opts.app_label}_{self.opts.model_name}_add")
        params = request.GET.copy()
        params["archived"] = "1"
        archive_url = f"{request.path}?{params.urlencode()}"

        response.context_data.update(
            {
                "individual_course_rows": row_items,
                "individual_add_url": add_url,
                "individual_archive_url": archive_url,
                "individual_search_query": (request.GET.get("q") or "").strip(),
                "individual_is_archive": request.GET.get("archived") == "1",
                "course_create_url": reverse(f"{crm_admin_site.name}:course_quick_create"),
                "course_status_choices": _course_status_choices(),
                "course_subject_choices": list(
                    self.get_queryset(request)
                    .exclude(subject="")
                    .values_list("subject", flat=True)
                    .distinct()
                    .order_by("subject")
                ),
            }
        )
        return response


@admin.register(Lead, site=crm_admin_site)
class LeadAdmin(RoleRestrictedAdminMixin, OrganizationFilterMixin, ArchiveAdminMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор", "Менеджер"}
    form = LeadAdminForm
    change_list_template = "admin/leads_changelist.html"
    list_display = ("full_name", "phone_number", "source", "message_short", "status", "created_at")
    list_filter = ("status", ArchiveFilter)
    search_fields = ("full_name", "phone_number", "email", "message", "conversation_log")
    actions = (archive_selected, unarchive_selected)
    readonly_fields = ("created_at", "source", "session_key", "conversation_log")
    fieldsets = (
        (
            None,
            {
                "fields": (
                    "organization",
                    "assignee",
                    "first_name",
                    "last_name",
                    "middle_name",
                    "full_name",
                    "phone_number",
                    "extra_phone",
                    "email",
                    "telegram_nick",
                    "birth_date",
                    "due_date",
                    "from_where",
                    "channel",
                    "source",
                    "status",
                    "subject_interest",
                    "interested_courses",
                    "labels",
                    "comment",
                    "lost_reason",
                    "created_at",
                )
            },
        ),
        (
            "Чат",
            {
                "fields": (
                    "session_key",
                    "message",
                    "bot_reply",
                    "conversation_log",
                )
            },
        ),
    )

    def message_short(self, obj: Lead):
        if not obj.message:
            return "—"
        return obj.message[:50] + "…" if len(obj.message) > 50 else obj.message
    message_short.short_description = "Вопрос"

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path("create/", self.admin_site.admin_view(self.lead_quick_create), name="settings_lead_quick_create"),
            path("upload/", self.admin_site.admin_view(self.lead_upload), name="settings_lead_upload"),
            path("upload/meta/", self.admin_site.admin_view(self.lead_upload_meta), name="settings_lead_upload_meta"),
            path("upload/template.xlsx", self.admin_site.admin_view(self.lead_upload_template), name="settings_lead_upload_template"),
            path(
                "upload/meta-template.csv",
                self.admin_site.admin_view(self.lead_upload_meta_template),
                name="settings_lead_upload_meta_template",
            ),
            path("export/", self.admin_site.admin_view(self.lead_export_fields), name="settings_lead_export_fields"),
            path("mailing-list/", self.admin_site.admin_view(self.lead_mailing_list), name="settings_lead_mailing_list"),
            path("archived/", self.admin_site.admin_view(self.lead_archive_page), name="settings_lead_archive"),
            path("duplicates/", self.admin_site.admin_view(self.lead_duplicates), name="settings_lead_duplicates"),
            path("reports/", self.admin_site.admin_view(self.lead_reports), name="settings_lead_reports"),
            path("sales-report/", self.admin_site.admin_view(self.lead_sales_report), name="settings_lead_sales_report"),
            path(
                "<int:lead_id>/status/",
                self.admin_site.admin_view(self.lead_status_update),
                name="settings_lead_status_update",
            ),
        ]
        return custom + urls

    @staticmethod
    def _lead_status_label(status: str) -> str:
        return LEAD_STATUS_LABELS.get(status, status or "—")

    @staticmethod
    def _lead_source_label(source: str) -> str:
        source = (source or "manual").strip()
        labels = {
            "manual": "Ручной",
            "website": "Сайт",
            "website_chat": "Чат сайта",
            "instagram": "Instagram",
            "telegram": "Telegram",
            "whatsapp": "WhatsApp",
        }
        return labels.get(source, source.replace("_", " ").title())

    @staticmethod
    def _lead_source_mark(source: str) -> str:
        source = (source or "manual").strip().lower()
        if "chat" in source:
            return "C"
        if "instagram" in source:
            return "I"
        if "telegram" in source:
            return "T"
        if "whatsapp" in source:
            return "W"
        if "site" in source or "web" in source:
            return "A"
        return "A"

    @staticmethod
    def _lead_column_values(column_key: str) -> tuple[str, ...]:
        for column in LEAD_BOARD_COLUMNS:
            if column["key"] == column_key:
                return tuple(column["values"])
        if column_key in LEAD_STATUS_LABELS:
            return (column_key,)
        return ()

    @staticmethod
    def _lead_status_value(raw: str | None) -> str:
        value = (raw or "").strip()
        if not value:
            return Lead.Status.NEW
        if value in LEAD_STATUS_LABELS:
            return value
        lowered = value.casefold()
        for key, label in LEAD_STATUS_CHOICES:
            if lowered in {key.casefold(), label.casefold()}:
                return key
        for column in LEAD_BOARD_COLUMNS:
            if lowered == column["title"].casefold():
                return column["values"][0]
        return Lead.Status.NEW

    @staticmethod
    def _lead_clip(value: str | None, length: int) -> str:
        return (value or "").strip()[:length]

    @staticmethod
    def _lead_full_name(first_name: str, last_name: str, middle_name: str, fallback: str = "") -> str:
        full_name = " ".join(part for part in (last_name, first_name, middle_name) if part).strip()
        return full_name or (fallback or "").strip() or "Посетитель сайта"

    def _lead_assignee_from_value(self, raw: str | None):
        value = (raw or "").strip()
        if not value:
            return None
        qs = User.objects.filter(is_active=True)
        if value.isdigit():
            return qs.filter(pk=int(value)).first()
        lowered = value.casefold()
        for user in qs:
            names = {
                user.username.casefold(),
                (user.get_full_name() or "").strip().casefold(),
                f"{user.last_name} {user.first_name}".strip().casefold(),
                f"{user.first_name} {user.last_name}".strip().casefold(),
            }
            if lowered in names:
                return user
        return None

    def _lead_people(self):
        return User.objects.filter(is_active=True, is_staff=True).order_by("last_name", "first_name", "username")

    def _lead_courses(self):
        return Cursues.objects.filter(is_archived=False).order_by("title")

    def _lead_base_queryset(self, request, *, include_archived: bool = False):
        qs = Lead.objects.select_related("assignee", "organization").all()
        org_id = request.session.get("current_org_id")
        if org_id:
            qs = qs.filter(organization_id=org_id)
        if not include_archived:
            qs = qs.filter(is_archived=False)
        return qs

    def _lead_payload_from_post(self, request) -> dict:
        return {
            "first_name": request.POST.get("first_name"),
            "last_name": request.POST.get("last_name"),
            "middle_name": request.POST.get("middle_name"),
            "full_name": request.POST.get("full_name"),
            "due_date": request.POST.get("due_date"),
            "phone_number": request.POST.get("phone_number"),
            "extra_phone": request.POST.get("extra_phone"),
            "email": request.POST.get("email"),
            "telegram_nick": request.POST.get("telegram_nick"),
            "birth_date": request.POST.get("birth_date"),
            "from_where": request.POST.get("from_where"),
            "assignee": request.POST.get("assignee"),
            "channel": request.POST.get("channel"),
            "status": request.POST.get("status"),
            "subject_interest": request.POST.get("subject_interest"),
            "interested_courses": request.POST.get("interested_courses"),
            "labels": request.POST.get("labels"),
            "comment": request.POST.get("comment"),
        }

    def _lead_create_from_payload(self, request, data: dict) -> Lead:
        first_name = self._lead_clip(data.get("first_name"), 150)
        last_name = self._lead_clip(data.get("last_name"), 150)
        middle_name = self._lead_clip(data.get("middle_name"), 150)
        full_name = self._lead_full_name(first_name, last_name, middle_name, data.get("full_name"))
        channel = self._lead_clip(data.get("channel"), 64)
        source = self._lead_clip(data.get("source") or channel or "manual", 32) or "manual"
        assignee = self._lead_assignee_from_value(data.get("assignee"))
        org_id = request.session.get("current_org_id")

        lead = Lead.objects.create(
            organization_id=org_id or None,
            first_name=first_name,
            last_name=last_name,
            middle_name=middle_name,
            full_name=self._lead_clip(full_name, 255),
            phone_number=self._lead_clip(data.get("phone_number"), 64),
            extra_phone=self._lead_clip(data.get("extra_phone"), 64),
            email=self._lead_clip(data.get("email"), 255),
            telegram_nick=self._lead_clip(data.get("telegram_nick"), 100),
            birth_date=parse_date((data.get("birth_date") or "").strip()),
            due_date=parse_date((data.get("due_date") or "").strip()),
            from_where=self._lead_clip(data.get("from_where"), 255),
            assignee=assignee,
            channel=channel,
            source=source,
            status=self._lead_status_value(data.get("status")),
            subject_interest=self._lead_clip(data.get("subject_interest"), 255),
            interested_courses=(data.get("interested_courses") or "").strip(),
            labels=self._lead_clip(data.get("labels"), 255),
            comment=(data.get("comment") or "").strip(),
            message=(data.get("comment") or "").strip(),
        )
        return lead

    def _lead_export_value(self, lead: Lead, field: str) -> str:
        if field == "id":
            return str(lead.pk)
        if field == "status":
            return self._lead_status_label(lead.status)
        if field == "assignee":
            if not lead.assignee:
                return ""
            return lead.assignee.get_full_name() or lead.assignee.username
        if field == "created_at":
            return timezone.localtime(lead.created_at).strftime("%d.%m.%Y %H:%M")
        if field == "archived_at":
            return timezone.localtime(lead.archived_at).strftime("%d.%m.%Y %H:%M") if lead.archived_at else ""
        value = getattr(lead, field, "")
        if isinstance(value, date):
            return value.strftime("%d.%m.%Y")
        return str(value or "")

    def _lead_rows_from_file(self, file_obj):
        name = (getattr(file_obj, "name", "") or "").lower()
        if name.endswith((".xlsx", ".xlsm")):
            workbook = load_workbook(file_obj, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(cell or "").strip() for cell in rows[0]]
            data_rows = []
            for row in rows[1:]:
                data_rows.append({headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]})
            return data_rows

        text = file_obj.read().decode("utf-8-sig")
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        return list(csv.DictReader(io.StringIO(text), dialect=dialect))

    def _lead_payload_from_row(self, row: dict, mapping: dict | None = None) -> dict:
        def pick(*names):
            if mapping:
                column = mapping.get(names[0], "")
                if column:
                    return row.get(column, "")
            for name in names:
                if name in row:
                    return row.get(name, "")
            return ""

        return {
            "full_name": pick("full_name", "ФИО", "фио", "полное_имя", "name"),
            "first_name": pick("first_name", "Имя", "имя"),
            "last_name": pick("last_name", "Фамилия", "фамилия"),
            "middle_name": pick("middle_name", "Отчество", "отчество"),
            "phone_number": pick("phone_number", "Номер телефона", "Телефон", "номер_телефона"),
            "extra_phone": pick("extra_phone", "Дополнительный номер"),
            "email": pick("email", "Адрес электронной почты", "Email", "эл_адрес"),
            "comment": pick("comment", "Комментарий", "campaign_name"),
            "channel": pick("channel", "Канал", "platform"),
            "source": pick("source", "Источник", "platform"),
            "status": pick("status", "Статус"),
            "subject_interest": pick("subject_interest", "Интересуется предметом"),
            "interested_courses": pick("interested_courses", "Интересующие курсы", "Курс"),
            "assignee": pick("assignee", "Ответственный", "Ответственные"),
        }

    def _apply_board_filters(self, request, qs):
        q = (request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(full_name__icontains=q)
                | Q(phone_number__icontains=q)
                | Q(email__icontains=q)
                | Q(message__icontains=q)
                | Q(conversation_log__icontains=q)
            )

        source = (request.GET.get("source") or "").strip()
        if source:
            qs = qs.filter(source=source)

        status = (request.GET.get("status") or "").strip()
        status_values = self._lead_column_values(status)
        if status_values:
            qs = qs.filter(status__in=status_values)

        created_from = parse_date((request.GET.get("created_from") or "").strip())
        if created_from:
            qs = qs.filter(created_at__date__gte=created_from)

        created_to = parse_date((request.GET.get("created_to") or "").strip())
        if created_to:
            qs = qs.filter(created_at__date__lte=created_to)

        return qs

    def _lead_export_csv(self, request, selected_fields: list[str] | None = None):
        allowed = {key for key, _label, _default in LEAD_EXPORT_FIELDS}
        selected_fields = [field for field in (selected_fields or request.GET.getlist("fields")) if field in allowed]
        if not selected_fields:
            selected_fields = [key for key, _label, default in LEAD_EXPORT_FIELDS if default]
        labels = dict((key, label) for key, label, _default in LEAD_EXPORT_FIELDS)

        qs = self._apply_board_filters(request, self.get_queryset(request)).select_related("assignee").order_by("-created_at")
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="leads.csv"'
        response.write("\ufeff")
        writer = csv.writer(response)
        writer.writerow([labels[field] for field in selected_fields])
        for lead in qs:
            writer.writerow([self._lead_export_value(lead, field) for field in selected_fields])
        return response

    def _lead_card(self, request, lead: Lead) -> dict:
        created_at = timezone.localtime(lead.created_at)
        column_key = LEAD_STATUS_TO_COLUMN.get(lead.status, "consultation")
        column_meta = next((column for column in LEAD_BOARD_COLUMNS if column["key"] == column_key), LEAD_BOARD_COLUMNS[0])
        user_name = (
            (lead.assignee.get_full_name() or lead.assignee.username)
            if lead.assignee
            else (request.user.get_full_name() or request.user.get_username())
        )
        message = (lead.comment or lead.message or lead.bot_reply or lead.conversation_log or "").strip()
        phone_or_email = lead.phone_number or lead.email
        return {
            "id": lead.pk,
            "name": lead.full_name or "Посетитель сайта",
            "phone_or_email": phone_or_email,
            "message": message[:110] + "..." if len(message) > 110 else message,
            "status": lead.status,
            "status_label": self._lead_status_label(lead.status),
            "badge": column_meta["badge"],
            "tone": column_meta["tone"],
            "source": self._lead_source_label(lead.source),
            "source_mark": self._lead_source_mark(lead.source),
            "created_date": created_at.date(),
            "next_contact_date": lead.due_date or (created_at.date() + timedelta(days=10)),
            "responsible": user_name,
            "change_url": reverse(f"{self.admin_site.name}:settings_lead_change", args=[lead.pk]),
            "status_url": reverse(f"{self.admin_site.name}:settings_lead_status_update", args=[lead.pk]),
        }

    def lead_quick_create(self, request):
        if request.method != "POST":
            return redirect(reverse(f"{self.admin_site.name}:settings_lead_changelist"))
        try:
            self._lead_create_from_payload(request, self._lead_payload_from_post(request))
            messages.success(request, "Лид добавлен.")
        except Exception as exc:
            messages.error(request, f"Не удалось добавить лид: {exc}")
        return redirect(reverse(f"{self.admin_site.name}:settings_lead_changelist"))

    def lead_upload_template(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Leads"
        sheet.append(["ФИО", "Телефон", "Email", "Комментарий", "Канал", "Статус"])
        sheet.append(["Иван Иванов", "+996555000000", "lead@example.com", "Интересуется курсом", "instagram", "Консультация"])
        buf = io.BytesIO()
        workbook.save(buf)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="leads-template.xlsx"'
        return response

    def lead_upload_meta_template(self, request):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="leads-meta-template.csv"'
        response.write("\ufeff")
        writer = csv.writer(response)
        writer.writerow(["полное_имя", "эл_адрес", "номер_телефона", "campaign_name", "platform"])
        writer.writerow(["Ivan Lead", "lead@example.com", "+996555000000", "IELTS Campaign", "facebook"])
        return response

    def lead_upload(self, request):
        if request.method == "POST":
            file_obj = request.FILES.get("file")
            if not file_obj:
                messages.error(request, "Выберите файл.")
                return redirect(reverse(f"{self.admin_site.name}:settings_lead_upload"))
            try:
                created = 0
                for row in self._lead_rows_from_file(file_obj):
                    payload = self._lead_payload_from_row(row)
                    if any(str(value or "").strip() for value in payload.values()):
                        self._lead_create_from_payload(request, payload)
                        created += 1
                messages.success(request, f"Загружено лидов: {created}.")
                return redirect(reverse(f"{self.admin_site.name}:settings_lead_changelist"))
            except Exception as exc:
                messages.error(request, f"Не удалось загрузить файл: {exc}")
                return redirect(reverse(f"{self.admin_site.name}:settings_lead_upload"))

        context = {
            **self.admin_site.each_context(request),
            "title": "Загрузить список",
            "opts": self.model._meta,
            "template_url": reverse(f"{self.admin_site.name}:settings_lead_upload_template"),
            "lead_list_url": reverse(f"{self.admin_site.name}:settings_lead_changelist"),
        }
        return TemplateResponse(request, "admin/leads_upload.html", context)

    def lead_upload_meta(self, request):
        default_mapping = {
            "full_name": "полное_имя",
            "email": "эл_адрес",
            "phone_number": "номер_телефона",
            "comment": "campaign_name",
            "channel": "platform",
        }
        if request.method == "POST":
            file_obj = request.FILES.get("file")
            if not file_obj:
                messages.error(request, "Выберите CSV-файл.")
                return redirect(reverse(f"{self.admin_site.name}:settings_lead_upload_meta"))
            mapping = {
                "full_name": (request.POST.get("map_full_name") or "").strip(),
                "email": (request.POST.get("map_email") or "").strip(),
                "phone_number": (request.POST.get("map_phone_number") or "").strip(),
                "comment": (request.POST.get("map_comment") or "").strip(),
                "channel": (request.POST.get("map_channel") or "").strip(),
            }
            course = (request.POST.get("course") or "").strip()
            assignee = (request.POST.get("assignee") or "").strip()
            try:
                created = 0
                for row in self._lead_rows_from_file(file_obj):
                    payload = self._lead_payload_from_row(row, mapping)
                    payload["interested_courses"] = course
                    payload["assignee"] = assignee
                    if any(str(value or "").strip() for value in payload.values()):
                        self._lead_create_from_payload(request, payload)
                        created += 1
                messages.success(request, f"Загружено лидов Meta: {created}.")
                return redirect(reverse(f"{self.admin_site.name}:settings_lead_changelist"))
            except Exception as exc:
                messages.error(request, f"Не удалось загрузить Meta-файл: {exc}")
                return redirect(reverse(f"{self.admin_site.name}:settings_lead_upload_meta"))

        context = {
            **self.admin_site.each_context(request),
            "title": "Загрузить список: Meta",
            "opts": self.model._meta,
            "mapping": default_mapping,
            "lead_list_url": reverse(f"{self.admin_site.name}:settings_lead_changelist"),
            "template_url": reverse(f"{self.admin_site.name}:settings_lead_upload_meta_template"),
            "people": self._lead_people(),
            "courses": self._lead_courses(),
        }
        return TemplateResponse(request, "admin/leads_upload_meta.html", context)

    def lead_export_fields(self, request):
        if request.method == "POST":
            return self._lead_export_csv(request, selected_fields=request.POST.getlist("fields"))
        context = {
            **self.admin_site.each_context(request),
            "title": "Выберите поля для экспорта",
            "opts": self.model._meta,
            "lead_list_url": reverse(f"{self.admin_site.name}:settings_lead_changelist"),
            "lead_selected_count": self._apply_board_filters(request, self.get_queryset(request)).count(),
            "export_fields": [
                {"key": key, "label": label, "default": default}
                for key, label, default in LEAD_EXPORT_FIELDS
            ],
        }
        return TemplateResponse(request, "admin/leads_export_fields.html", context)

    def lead_mailing_list(self, request):
        context = {
            **self.admin_site.each_context(request),
            "title": "Рассылка",
            "opts": self.model._meta,
            "lead_list_url": reverse(f"{self.admin_site.name}:settings_lead_changelist"),
        }
        return TemplateResponse(request, "admin/leads_mailing_list.html", context)

    def lead_reports(self, request):
        context = {
            **self.admin_site.each_context(request),
            "title": "Сводные отчеты по лидам",
            "opts": self.model._meta,
            "lead_list_url": reverse(f"{self.admin_site.name}:settings_lead_changelist"),
        }
        return TemplateResponse(request, "admin/leads_reports.html", context)

    def lead_archive_page(self, request):
        q = (request.GET.get("q") or "").strip()
        qs = self._lead_base_queryset(request, include_archived=True).filter(is_archived=True)
        if q:
            qs = qs.filter(
                Q(full_name__icontains=q)
                | Q(first_name__icontains=q)
                | Q(last_name__icontains=q)
                | Q(phone_number__icontains=q)
                | Q(email__icontains=q)
                | Q(lost_reason__icontains=q)
            )
        rows = qs.order_by("-archived_at", "-created_at")[:300]
        context = {
            **self.admin_site.each_context(request),
            "title": "Архив лидов",
            "opts": self.model._meta,
            "lead_list_url": reverse(f"{self.admin_site.name}:settings_lead_changelist"),
            "lead_archive_rows": rows,
            "lead_archive_total": qs.count(),
            "lead_archive_query": q,
        }
        return TemplateResponse(request, "admin/leads_archive.html", context)

    def lead_duplicates(self, request):
        groups: dict[str, dict] = {}
        qs = self._lead_base_queryset(request).order_by("-created_at")
        for lead in qs:
            keys = []
            phone_key = "".join(ch for ch in (lead.phone_number or "") if ch.isdigit())
            email_key = (lead.email or "").strip().casefold()
            if len(phone_key) >= 5:
                keys.append(("phone", phone_key, lead.phone_number))
            if email_key:
                keys.append(("email", email_key, lead.email))
            for kind, value, label in keys:
                key = f"{kind}:{value}"
                groups.setdefault(
                    key,
                    {
                        "label": f"{'Телефон' if kind == 'phone' else 'Email'}: {label}",
                        "leads": [],
                    },
                )["leads"].append(lead)

        duplicate_groups = [
            group for group in groups.values()
            if len(group["leads"]) > 1
        ]
        duplicate_groups.sort(key=lambda group: len(group["leads"]), reverse=True)
        context = {
            **self.admin_site.each_context(request),
            "title": "Дубликаты",
            "opts": self.model._meta,
            "lead_list_url": reverse(f"{self.admin_site.name}:settings_lead_changelist"),
            "lead_duplicate_groups": duplicate_groups[:80],
            "lead_duplicate_total": sum(len(group["leads"]) for group in duplicate_groups),
        }
        return TemplateResponse(request, "admin/leads_duplicates.html", context)

    def lead_sales_report(self, request):
        today = timezone.localdate()
        month_start = today.replace(day=1)
        date_from = parse_date((request.GET.get("date_from") or "").strip()) or month_start
        date_to = parse_date((request.GET.get("date_to") or "").strip()) or today
        channel = (request.GET.get("channel") or "").strip()

        qs = self._lead_base_queryset(request, include_archived=True).filter(
            status=Lead.Status.WON,
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
        )
        if channel:
            qs = qs.filter(channel=channel)

        rows = []
        grouped = (
            qs.values("assignee_id", "assignee__first_name", "assignee__last_name", "assignee__username")
            .annotate(total=Count("id"))
            .order_by("assignee__last_name", "assignee__first_name", "assignee__username")
        )
        for row in grouped:
            name = " ".join(
                part for part in (row["assignee__last_name"], row["assignee__first_name"]) if part
            ).strip()
            rows.append({"name": name or row["assignee__username"] or "Без ответственного", "total": row["total"]})

        channel_choices = (
            self._lead_base_queryset(request, include_archived=True)
            .exclude(channel="")
            .values_list("channel", flat=True)
            .distinct()
            .order_by("channel")
        )
        context = {
            **self.admin_site.each_context(request),
            "title": "Отчет по продажам",
            "opts": self.model._meta,
            "lead_list_url": reverse(f"{self.admin_site.name}:settings_lead_changelist"),
            "lead_sales_rows": rows,
            "lead_sales_total": sum(row["total"] for row in rows),
            "lead_sales_date_from": date_from,
            "lead_sales_date_to": date_to,
            "lead_sales_channel": channel,
            "lead_sales_channels": channel_choices,
        }
        return TemplateResponse(request, "admin/leads_sales_report.html", context)

    def lead_status_update(self, request, lead_id: int):
        if request.method != "POST":
            return JsonResponse({"ok": False, "error": "method not allowed"}, status=405)
        if not self.has_change_permission(request):
            return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

        status = (request.POST.get("status") or "").strip()
        if status not in LEAD_STATUS_LABELS:
            return JsonResponse({"ok": False, "error": "invalid status"}, status=400)

        lead = get_object_or_404(self.get_queryset(request), pk=lead_id)
        lead.status = status
        lead.save(update_fields=["status"])

        column_key = LEAD_STATUS_TO_COLUMN.get(status, "consultation")
        column_meta = next((column for column in LEAD_BOARD_COLUMNS if column["key"] == column_key), LEAD_BOARD_COLUMNS[0])
        return JsonResponse(
            {
                "ok": True,
                "status": status,
                "column": column_key,
                "badge": column_meta["badge"],
                "tone": column_meta["tone"],
            }
        )

    def changelist_view(self, request, extra_context=None):
        if request.GET.get("export") == "csv":
            return self._lead_export_csv(request)

        if not self.has_view_or_change_permission(request):
            return HttpResponseForbidden("Forbidden")

        base_qs = self.get_queryset(request)
        filtered_qs = self._apply_board_filters(request, base_qs).order_by("-created_at")
        source_choices = (
            base_qs.exclude(source="")
            .values_list("source", flat=True)
            .distinct()
            .order_by("source")
        )

        buckets: dict[str, list[dict]] = {column["key"]: [] for column in LEAD_BOARD_COLUMNS}
        for lead in filtered_qs:
            card = self._lead_card(request, lead)
            buckets.setdefault(LEAD_STATUS_TO_COLUMN.get(card["status"], "consultation"), []).append(card)

        columns = []
        for column in LEAD_BOARD_COLUMNS:
            cards = buckets.get(column["key"], [])
            columns.append(
                {
                    **column,
                    "target_status": column["values"][0],
                    "count": len(cards),
                    "cards": cards,
                }
            )

        params = request.GET.copy()
        params["export"] = "csv"
        export_url = f"{request.path}?{params.urlencode()}"
        export_fields_params = request.GET.copy()
        export_fields_params.pop("export", None)
        export_fields_url = reverse(f"{self.admin_site.name}:settings_lead_export_fields")
        if export_fields_params:
            export_fields_url = f"{export_fields_url}?{export_fields_params.urlencode()}"

        filter_values = {
            "q": (request.GET.get("q") or "").strip(),
            "status": (request.GET.get("status") or "").strip(),
            "source": (request.GET.get("source") or "").strip(),
            "created_from": (request.GET.get("created_from") or "").strip(),
            "created_to": (request.GET.get("created_to") or "").strip(),
            "archived": (request.GET.get("archived") or "").strip(),
        }
        active_filters_count = sum(1 for key, value in filter_values.items() if key != "archived" and value)
        if filter_values["archived"] == "1":
            active_filters_count += 1

        try:
            tasks_url = reverse(f"{self.admin_site.name}:settings_task_changelist")
        except Exception:
            tasks_url = "#"
        try:
            lead_content_type = ContentType.objects.get_for_model(Lead)
            history_url = (
                reverse(f"{self.admin_site.name}:admin_logentry_changelist")
                + f"?content_type__id__exact={lead_content_type.pk}"
            )
        except Exception:
            history_url = reverse(f"{self.admin_site.name}:index")

        context = {
            **self.admin_site.each_context(request),
            "title": "Лиды",
            "opts": self.model._meta,
            "media": self.media,
            "lead_board_columns": columns,
            "lead_total": sum(column["count"] for column in columns),
            "lead_filter_values": filter_values,
            "lead_filter_statuses": LEAD_BOARD_COLUMNS,
            "lead_source_choices": [
                {"value": source, "label": self._lead_source_label(source)}
                for source in source_choices
            ],
            "lead_active_filters_count": active_filters_count,
            "lead_filters_open": active_filters_count > 0,
            "lead_add_url": reverse(f"{self.admin_site.name}:settings_lead_quick_create"),
            "lead_change_add_url": reverse(f"{self.admin_site.name}:settings_lead_add"),
            "lead_upload_url": reverse(f"{self.admin_site.name}:settings_lead_upload"),
            "lead_upload_meta_url": reverse(f"{self.admin_site.name}:settings_lead_upload_meta"),
            "lead_export_url": export_fields_url,
            "lead_tasks_url": tasks_url,
            "lead_forms_url": reverse(f"{self.admin_site.name}:settings_index"),
            "lead_mailing_url": reverse(f"{self.admin_site.name}:settings_lead_mailing_list"),
            "lead_archive_url": reverse(f"{self.admin_site.name}:settings_lead_archive"),
            "lead_history_url": history_url,
            "lead_duplicates_url": reverse(f"{self.admin_site.name}:settings_lead_duplicates"),
            "lead_reports_url": reverse(f"{self.admin_site.name}:settings_lead_reports"),
            "lead_sales_report_url": reverse(f"{self.admin_site.name}:settings_lead_sales_report"),
            "lead_reset_url": request.path,
            "lead_status_choices": LEAD_STATUS_CHOICES,
            "lead_people": self._lead_people(),
            "lead_courses": self._lead_courses(),
            "is_archive": filter_values["archived"] == "1",
            "has_add_permission": self.has_add_permission(request),
        }
        if extra_context:
            context.update(extra_context)
        return TemplateResponse(request, self.change_list_template, context)


@admin.register(Payment, site=crm_admin_site)
class PaymentAdmin(RoleRestrictedAdminMixin, OrganizationFilterMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор", "Менеджер"}
    list_display = ("student", "course", "amount", "created_at")
    list_filter = ("course",)
    search_fields = ("student__user__username", "student__user__phone_number")
    autocomplete_fields = ("student", "course")


@admin.register(Salary, site=crm_admin_site)
class SalaryAdmin(RoleRestrictedAdminMixin, OrganizationFilterMixin, admin.ModelAdmin):
    denied_roles = {"Менеджер", "Ментор"}
    list_display = ("mentor", "course", "amount", "comment_short", "created_at")
    list_filter = ("created_at",)
    search_fields = ("mentor__user__first_name", "mentor__user__last_name", "mentor__user__username", "comment")
    autocomplete_fields = ("mentor", "course")
    fields = ("mentor", "course", "amount", "comment", "created_at")
    readonly_fields = ("created_at",)

    def comment_short(self, obj: Salary):
        return obj.comment[:50] + "…" if len(obj.comment) > 50 else obj.comment
    comment_short.short_description = "Комментарий"


@admin.register(Task, site=crm_admin_site)
class TaskAdmin(RoleRestrictedAdminMixin, OrganizationFilterMixin, ArchiveAdminMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор", "Менеджер", "Ментор"}
    list_display = ("title", "due_date", "is_done", "created_at")
    list_filter = ("is_done", ArchiveFilter)
    search_fields = ("title",)
    actions = (archive_selected, unarchive_selected)


class FinanceStatusFilter(admin.SimpleListFilter):
    title = "Финансы"
    parameter_name = "finance"

    def lookups(self, request, model_admin):
        return (
            ("paid", "Оплачено"),
            ("debt", "Должники"),
            ("partial", "Частично"),
        )

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset

        if value == "paid":
            return queryset.extra(
                where=[
                    "(SELECT COALESCE(SUM(amount), 0) FROM settings_payment p WHERE p.student_id = settings_enrollment.student_id AND p.course_id = settings_enrollment.course_id) >= settings_enrollment.tuition_amount"
                ]
            )
        if value == "debt":
            return queryset.extra(
                where=["(SELECT COALESCE(SUM(amount), 0) FROM settings_payment p WHERE p.student_id = settings_enrollment.student_id AND p.course_id = settings_enrollment.course_id) < settings_enrollment.tuition_amount"]
            )
        if value == "partial":
            return queryset.extra(
                where=[
                    "(SELECT COALESCE(SUM(amount), 0) FROM settings_payment p WHERE p.student_id = settings_enrollment.student_id AND p.course_id = settings_enrollment.course_id) > 0",
                    "(SELECT COALESCE(SUM(amount), 0) FROM settings_payment p WHERE p.student_id = settings_enrollment.student_id AND p.course_id = settings_enrollment.course_id) < settings_enrollment.tuition_amount",
                ]
            )
        return queryset


class EnrollmentBaseAdmin(admin.ModelAdmin):
    list_display = ("student", "course", "tuition_amount", "paid_total_display", "debt_display", "status_display")
    list_filter = ("course", FinanceStatusFilter)
    search_fields = ("student__user__username", "student__user__phone_number", "course__title")
    autocomplete_fields = ("student", "course")

    def paid_total_display(self, obj: Enrollment):
        total = (
            Payment.objects.filter(student=obj.student, course=obj.course).aggregate(total=Sum("amount"))["total"]
            or 0
        )
        return total

    paid_total_display.short_description = "Оплачено (с.)"

    def debt_display(self, obj: Enrollment):
        debt = obj.debt
        return format_html(
            '<span style="font-weight:700;color:{};">{}</span>',
            "#ef4444" if debt > 0 else "#10b981",
            debt,
        )

    debt_display.short_description = "Долг (с.)"

    def status_display(self, obj: Enrollment):
        debt = obj.debt
        if debt <= 0:
            return format_html('<span class="badge badge-success">Оплачено</span>')
        if obj.paid_total > 0:
            return format_html('<span class="badge badge-warning">Частично</span>')
        return format_html('<span class="badge badge-danger">Долг</span>')

    status_display.short_description = "Статус"


@admin.register(StudentPayments, site=crm_admin_site)
class StudentPaymentsAdmin(RoleRestrictedAdminMixin, EnrollmentBaseAdmin):
    allowed_roles = {"Администратор", "Менеджер"}
    def changelist_view(self, request, extra_context=None):
        if "finance" not in request.GET:
            mutable = request.GET.copy()
            mutable["finance"] = "paid"
            request.GET = mutable
            request.META["QUERY_STRING"] = mutable.urlencode()
        return super().changelist_view(request, extra_context=extra_context)


@admin.register(TuitionPayment, site=crm_admin_site)
class TuitionPaymentAdmin(RoleRestrictedAdminMixin, OrganizationFilterMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор", "Менеджер"}
    change_list_template = "admin/tuition_payment_changelist.html"
    list_per_page = 20
    list_display = (
        "payment_id_display",
        "student_name_display",
        "pay_datetime_display",
        "method_display",
        "course_display",
        "amount_display",
        "comment_display",
        "actions_display",
    )
    list_display_links = None
    search_fields = ("student__user__username", "student__user__first_name", "student__user__last_name", "student__user__phone_number")
    autocomplete_fields = ("student", "course")
    ordering = ("-created_at",)
    actions = None

    @staticmethod
    def _get_int_param(request, key: str):
        params = getattr(request, "_crm_pay_params", request.GET)
        raw = (params.get(key) or "").strip()
        if not raw:
            return None
        try:
            return int(raw)
        except (TypeError, ValueError):
            return None

    def payment_id_display(self, obj: Payment):
        first_id = getattr(obj, "_crm_first_pay_id", None)
        is_first = first_id is not None and obj.pk == first_id
        color = "#10b981" if is_first else "#6b7280"
        return format_html('<span class="crm-pay-id" style="color:{};font-weight:700;">#{}</span>', color, obj.pk)

    payment_id_display.short_description = "№"

    def student_name_display(self, obj: Payment):
        u = obj.student.user
        name = u.get_full_name() or u.username
        return format_html('<span class="crm-pay-td-name">{}</span>', name)

    student_name_display.short_description = "ФИО"

    def pay_datetime_display(self, obj: Payment):
        if not obj.created_at:
            return "—"
        dt = timezone.localtime(obj.created_at)
        return dt.strftime("%d.%m.%Y %H:%M")

    pay_datetime_display.short_description = "Дата и время"
    pay_datetime_display.admin_order_field = "created_at"

    def method_display(self, obj: Payment):
        return obj.get_method_display()

    method_display.short_description = "Способ"
    method_display.admin_order_field = "method"

    def course_display(self, obj: Payment):
        title = obj.course.title if obj.course_id else "—"
        return format_html('<span class="crm-pay-td-course">{}</span>', title)

    course_display.short_description = "Курс"
    course_display.admin_order_field = "course__title"

    def amount_display(self, obj: Payment):
        return format_html(
            '<span class="crm-pay-td-amount">+ {} с.</span>',
            int(obj.amount) if obj.amount == int(obj.amount) else obj.amount,
        )

    amount_display.short_description = "Сумма"
    amount_display.admin_order_field = "amount"

    def comment_display(self, obj: Payment):
        if obj.is_voided:
            return format_html('<span class="crm-pay-td-comment" style="color:#ef4444;">Аннулирован</span>')
        text = (obj.description or "").strip()
        if not text:
            return format_html('<span class="crm-pay-td-comment">—</span>')
        return format_html('<span class="crm-pay-td-comment" title="{}">{}</span>', text, text[:20] + "..." if len(text) > 20 else text)

    comment_display.short_description = "Коммент"

    def actions_display(self, obj: Payment):
        return format_html(
            '<div class="crm-pay-actions-dd">'
            '<button type="button" class="crm-pay-actions-trigger" data-pay-dd-toggle aria-expanded="false" aria-haspopup="true" aria-label="Действия">⋮</button>'
            '<div class="crm-pay-actions-menu" role="menu" hidden>'
            '<button type="button" class="crm-pay-actions-item" data-pay-action="edit" data-pay-id="{}" role="menuitem">'
            '<i class="fas fa-edit" aria-hidden="true"></i> Редактировать</button>'
            '<button type="button" class="crm-pay-actions-item" data-pay-action="void" data-pay-id="{}" role="menuitem">'
            '<i class="fas fa-times-circle" aria-hidden="true"></i> Аннулировать</button>'
            '<button type="button" class="crm-pay-actions-item" data-pay-action="receipt" data-pay-id="{}" role="menuitem">'
            '<i class="fas fa-download" aria-hidden="true"></i> Выгрузить чек</button>'
            '<button type="button" class="crm-pay-actions-item" data-pay-action="attach" data-pay-id="{}" role="menuitem">'
            '<i class="fas fa-upload" aria-hidden="true"></i> Прикрепить квитанцию-подтверждение</button>'
            "</div></div>",
            obj.pk, obj.pk, obj.pk, obj.pk
        )

    actions_display.short_description = "Действия"

    def get_queryset(self, request):
        qs = super().get_queryset(request).select_related("student__user", "course").prefetch_related("course__mentors")

        course_id = self._get_int_param(request, "course")
        if course_id is not None:
            qs = qs.filter(course_id=course_id)

        mentor_id = self._get_int_param(request, "mentor")
        if mentor_id is not None:
            qs = qs.filter(course__mentors__id=mentor_id)

        params = getattr(request, "_crm_pay_params", request.GET)

        method = params.get("method")
        if method:
            qs = qs.filter(method=method)

        pay_id = self._get_int_param(request, "pay_id")
        if pay_id is not None:
            qs = qs.filter(id=pay_id)

        date_from = parse_date(params.get("date_from") or "")
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)

        date_to = parse_date(params.get("date_to") or "")
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        first = params.get("first")
        if first in ("0", "1"):
            first_created = (
                Payment.objects.filter(student_id=OuterRef("student_id"))
                .order_by("created_at")
                .values("created_at")[:1]
            )
            qs = qs.annotate(first_created_at=Subquery(first_created))
            if first == "1":
                qs = qs.filter(created_at=F("first_created_at"))
            else:
                qs = qs.exclude(created_at=F("first_created_at"))

        first_id_sub = (
            Payment.objects.filter(student_id=OuterRef("student_id"))
            .order_by("created_at", "id")
            .values("id")[:1]
        )
        qs = qs.annotate(_crm_first_pay_id=Subquery(first_id_sub))

        if mentor_id is not None:
            return qs.distinct()
        return qs

    def _tuition_export_xlsx(self, request):
        qs = self.get_queryset(request).order_by("-created_at")[:5000]
        wb = Workbook()
        ws = wb.active
        ws.title = "Оплата за учебу"
        headers = ("№", "ФИО", "Дата и время", "Способ", "Курс", "Сумма", "Комментарий")
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for p in qs:
            u = p.student.user
            name = u.get_full_name() or u.username
            dt_s = ""
            if p.created_at:
                dt_s = timezone.localtime(p.created_at).strftime("%d.%m.%Y %H:%M")
            ws.append(
                [
                    p.pk,
                    name,
                    dt_s,
                    p.get_method_display(),
                    p.course.title if p.course_id else "",
                    float(p.amount),
                    "",
                ]
            )
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="tuition-payments.xlsx"'
        return resp

    def _receipt_pdf(self, payment: Payment):
        return payment_receipt_pdf(payment)

    def changelist_view(self, request, extra_context=None):
        if request.method == "POST":
            action = (request.POST.get("action") or "").strip()
            pay_id = self._get_int_param(request, "pay_id")
            if pay_id is not None and action:
                payment = get_object_or_404(Payment, pk=pay_id)
                if action == "edit":
                    method = (request.POST.get("method") or "").strip()
                    description = (request.POST.get("description") or "").strip()
                    if method in dict(Payment.Method.choices):
                        payment.method = method
                    payment.description = description
                    payment.save(update_fields=["method", "description"])
                    return JsonResponse({"ok": True})
                if action == "void":
                    payment.is_voided = True
                    payment.save(update_fields=["is_voided"])
                    return JsonResponse({"ok": True})
                if action == "attach":
                    uploaded = request.FILES.get("receipt_file")
                    if uploaded:
                        if uploaded.size > 10 * 1024 * 1024:
                            return JsonResponse({"ok": False, "error": "Файл слишком большой (макс. 10 МБ)"}, status=400)
                        payment.receipt_file = uploaded
                        payment.save(update_fields=["receipt_file"])
                        return JsonResponse({"ok": True})
                    return JsonResponse({"ok": False, "error": "Файл не выбран"}, status=400)
                if action == "receipt":
                    content = self._receipt_pdf(payment)
                    resp = HttpResponse(content, content_type="application/pdf")
                    resp["Content-Disposition"] = f'attachment; filename="receipt_{payment.pk}.pdf"'
                    return resp

        request._crm_pay_params = request.GET.copy()
        mutable = request.GET.copy()
        for k in ("mentor", "first", "date_from", "date_to", "pay_id"):
            mutable.pop(k, None)
        request.GET = mutable
        request.META["QUERY_STRING"] = mutable.urlencode()

        if request._crm_pay_params.get("export") == "xlsx":
            return self._tuition_export_xlsx(request)

        if request._crm_pay_params.get("export") == "pdf":
            rows = []
            for p in self.get_queryset(request).order_by("-created_at")[:2000]:
                u = p.student.user
                name = u.get_full_name() or u.username
                rows.append(
                    [
                        str(p.id),
                        name,
                        p.created_at.strftime("%d.%m.%Y %H:%M") if p.created_at else "",
                        p.get_method_display() if hasattr(p, "get_method_display") else str(p.method),
                        str(p.course) if p.course_id else "",
                        str(p.amount),
                    ]
                )
            content = _simple_pdf_table(
                title="Tuition payments",
                headers=["ID", "ФИО", "Дата", "Способ оплаты", "Курс", "Сумма"],
                rows=rows,
            )
            resp = HttpResponse(content, content_type="application/pdf")
            resp["Content-Disposition"] = 'attachment; filename="tuition_payments.pdf"'
            return resp

        extra_context = extra_context or {}
        extra_context["request"] = request
        total_amount = self.get_queryset(request).aggregate(total=Sum("amount"))["total"] or 0
        extra_context["total_amount"] = total_amount
        extra_context["courses"] = Cursues.objects.filter(is_archived=False).order_by("title")
        extra_context["mentors"] = Mentor.objects.select_related("user").order_by("user__username")
        extra_context["methods"] = Payment.Method.choices
        extra_context["filters"] = {
            "course": request._crm_pay_params.get("course", ""),
            "mentor": request._crm_pay_params.get("mentor", ""),
            "method": request._crm_pay_params.get("method", ""),
            "first": request._crm_pay_params.get("first", ""),
            "date_from": request._crm_pay_params.get("date_from", ""),
            "date_to": request._crm_pay_params.get("date_to", ""),
            "pay_id": request._crm_pay_params.get("pay_id", ""),
        }
        extra_context["export_qs"] = request._crm_pay_params.urlencode()
        return super().changelist_view(request, extra_context=extra_context)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(CalendarEvent, site=crm_admin_site)
class CalendarEventAdmin(RoleRestrictedAdminMixin, OrganizationFilterMixin, ArchiveAdminMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор", "Менеджер", "Ментор"}
    list_display = ("title", "start_at", "end_at", "created_at")
    search_fields = ("title", "note")
    list_filter = ("start_at", ArchiveFilter)
    actions = (archive_selected, unarchive_selected)

    def changelist_view(self, request, extra_context=None):
        return redirect(reverse(f"{crm_admin_site.name}:calendar"))


@admin.register(Call, site=crm_admin_site)
class CallAdmin(RoleRestrictedAdminMixin, OrganizationFilterMixin, ArchiveAdminMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор", "Менеджер"}
    list_display = ("contact_name", "phone_number", "status", "created_at")
    list_filter = ("status", ArchiveFilter)
    search_fields = ("contact_name", "phone_number")
    actions = (archive_selected, unarchive_selected)


@admin.register(AccountingEntry, site=crm_admin_site)
class AccountingEntryAdmin(RoleRestrictedAdminMixin, OrganizationFilterMixin, ArchiveAdminMixin, admin.ModelAdmin):
    denied_roles = {"Менеджер", "Ментор"}
    change_list_template = "admin/accountingentry_changelist.html"
    list_display = ("entry_type", "title", "amount", "created_at")
    list_filter = ("entry_type", ArchiveFilter)
    search_fields = ("title",)
    actions = (archive_selected, unarchive_selected)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        params = getattr(request, "_crm_acc_params", request.GET)
        m = (params.get("m") or "").strip()
        if m:
            try:
                y_s, mo_s = m.split("-", 1)
                y = int(y_s)
                mo = int(mo_s)
                start = date(y, mo, 1)
                if mo == 12:
                    end = date(y + 1, 1, 1)
                else:
                    end = date(y, mo + 1, 1)
                qs = qs.filter(created_at__date__gte=start, created_at__date__lt=end)
            except Exception:
                pass

        q = (params.get("q") or "").strip()
        if q:
            qs = qs.filter(title__icontains=q)
        return qs

    def changelist_view(self, request, extra_context=None):
        request._crm_acc_params = request.GET.copy()
        mutable = request.GET.copy()
        mutable.pop("m", None)
        request.GET = mutable
        request.META["QUERY_STRING"] = mutable.urlencode()

        export = (request.GET.get("export") or "").strip().lower()
        if export == "pdf":
            rows = []
            for e in self.get_queryset(request).order_by("-created_at")[:2000]:
                rows.append(
                    [
                        str(e.entry_type),
                        str(e.title or ""),
                        str(e.amount),
                        e.created_at.strftime("%d.%m.%Y %H:%M") if e.created_at else "",
                    ]
                )
            content = _simple_pdf_table(
                title="Accounting",
                headers=["Тип", "Название", "Сумма", "Дата"],
                rows=rows,
            )
            resp = HttpResponse(content, content_type="application/pdf")
            resp["Content-Disposition"] = 'attachment; filename="accounting.pdf"'
            return resp

        extra_context = extra_context or {}
        extra_context["request"] = request
        params = request._crm_acc_params.copy()
        params.pop("export", None)
        extra_context["export_qs"] = params.urlencode()
        m = (request._crm_acc_params.get("m") or "").strip()
        extra_context["month"] = m or timezone.localdate().strftime("%Y-%m")
        extra_context["filters"] = {"q": (request._crm_acc_params.get("q") or "").strip()}

        qs = self.get_queryset(request)
        total_income = qs.filter(entry_type=AccountingEntry.Type.INCOME).aggregate(total=Sum("amount"))["total"] or 0
        total_expense = qs.filter(entry_type=AccountingEntry.Type.EXPENSE).aggregate(total=Sum("amount"))["total"] or 0
        extra_context["total_income"] = total_income
        extra_context["total_expense"] = total_expense
        extra_context["total_net"] = total_income - total_expense
        return super().changelist_view(request, extra_context=extra_context)


@admin.register(AccountingAccount, site=crm_admin_site)
class AccountingAccountAdmin(RoleRestrictedAdminMixin, OrganizationFilterMixin, admin.ModelAdmin):
    denied_roles = {"Менеджер", "Ментор"}
    list_display = ("id", "title", "created_at")
    search_fields = ("title",)


@admin.register(AccountingProject, site=crm_admin_site)
class AccountingProjectAdmin(RoleRestrictedAdminMixin, OrganizationFilterMixin, admin.ModelAdmin):
    denied_roles = {"Менеджер", "Ментор"}
    list_display = ("id", "title", "created_at")
    search_fields = ("title",)


@admin.register(AccountingCategory, site=crm_admin_site)
class AccountingCategoryAdmin(RoleRestrictedAdminMixin, OrganizationFilterMixin, admin.ModelAdmin):
    denied_roles = {"Менеджер", "Ментор"}
    list_display = ("id", "title", "created_at")
    search_fields = ("title",)

@admin.register(DebtorEnrollment, site=crm_admin_site)
class DebtorEnrollmentAdmin(RoleRestrictedAdminMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор", "Менеджер"}
    change_list_template = "admin/debtors_changelist.html"
    list_per_page = 20
    list_display = ("id", "student_name", "course", "tuition_amount", "paid_total_display", "debt_display")
    list_display_links = None
    ordering = ("-created_at",)
    actions = None

    @staticmethod
    def _get_int_param(request, key: str):
        params = getattr(request, "_crm_debt_params", request.GET)
        raw = (params.get(key) or "").strip()
        if not raw:
            return None
        try:
            return int(raw)
        except (TypeError, ValueError):
            return None

    def student_name(self, obj: Enrollment):
        u = obj.student.user
        return u.get_full_name() or u.username

    student_name.short_description = "ФИО"

    def get_queryset(self, request):
        qs = super().get_queryset(request).select_related("student__user", "course").prefetch_related("course__mentors")

        pay_sub = (
            Payment.objects.filter(student_id=OuterRef("student_id"), course_id=OuterRef("course_id"))
            .values("student_id")
            .annotate(total=Sum("amount"))
            .values("total")[:1]
        )
        qs = qs.annotate(
            paid_total_sum=Coalesce(
                Subquery(pay_sub),
                Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2)),
                output_field=models.DecimalField(max_digits=12, decimal_places=2),
            )
        ).annotate(
            debt_amount=F("tuition_amount") - F("paid_total_sum")
        ).filter(
            debt_amount__gt=0
        )

        course_id = self._get_int_param(request, "course")
        if course_id is not None:
            qs = qs.filter(course_id=course_id)

        mentor_id = self._get_int_param(request, "mentor")
        if mentor_id is not None:
            qs = qs.filter(course__mentors__id=mentor_id)

        student_q = (getattr(request, "_crm_debt_params", request.GET).get("q") or "").strip()
        if student_q:
            qs = qs.filter(
                models.Q(student__user__username__icontains=student_q)
                | models.Q(student__user__first_name__icontains=student_q)
                | models.Q(student__user__last_name__icontains=student_q)
                | models.Q(student__user__phone_number__icontains=student_q)
            )

        if mentor_id is not None:
            return qs.distinct()
        return qs

    def paid_total_display(self, obj: Enrollment):
        return obj.paid_total_sum

    paid_total_display.short_description = "Оплачено (с.)"

    def debt_display(self, obj: Enrollment):
        return obj.debt_amount

    debt_display.short_description = "Долг (с.)"

    def changelist_view(self, request, extra_context=None):
        request._crm_debt_params = request.GET.copy()
        mutable = request.GET.copy()
        for k in ("mentor", "q"):
            mutable.pop(k, None)
        request.GET = mutable
        request.META["QUERY_STRING"] = mutable.urlencode()

        if request._crm_debt_params.get("export") == "pdf":
            rows = []
            for e in self.get_queryset(request).order_by("-created_at")[:2000]:
                rows.append(
                    [
                        str(e.id),
                        self.student_name(e),
                        str(e.course) if e.course_id else "",
                        str(e.tuition_amount),
                        str(e.paid_total_sum),
                        str(e.debt_amount),
                    ]
                )
            content = _simple_pdf_table(
                title="Debtors",
                headers=["ID", "ФИО", "Курс", "К оплате", "Оплачено", "Долг"],
                rows=rows,
            )
            resp = HttpResponse(content, content_type="application/pdf")
            resp["Content-Disposition"] = 'attachment; filename="debtors.pdf"'
            return resp

        extra_context = extra_context or {}
        extra_context["request"] = request
        qs = self.get_queryset(request)
        extra_context["total_debt"] = qs.aggregate(total=Sum("debt_amount"))["total"] or 0
        extra_context["courses"] = Cursues.objects.filter(is_archived=False).order_by("title")
        extra_context["mentors"] = Mentor.objects.select_related("user").order_by("user__username")
        extra_context["filters"] = {
            "course": request._crm_debt_params.get("course", ""),
            "mentor": request._crm_debt_params.get("mentor", ""),
            "q": request._crm_debt_params.get("q", ""),
        }
        extra_context["export_qs"] = request._crm_debt_params.urlencode()
        return super().changelist_view(request, extra_context=extra_context)


class CRMLogEntryAdmin(RoleRestrictedAdminMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор", "Менеджер"}
    date_hierarchy = "action_time"
    ordering = ("-action_time",)
    list_display = ("action_time", "user", "content_type", "object_repr", "action_flag")
    list_filter = ("action_flag", "content_type")
    search_fields = ("object_repr", "change_message", "user__username")
    list_display_links = None

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(Lesson, site=crm_admin_site)
class LessonAdmin(RoleRestrictedAdminMixin, ArchiveAdminMixin, admin.ModelAdmin):
    allowed_roles = {"Администратор", "Менеджер"}
    list_display = ("title", "course", "mentor", "curriculum_module", "order", "is_archived")
    list_filter = ("course", "is_archived", "is_additional")
    search_fields = ("title", "description")
    autocomplete_fields = ("mentor", "course", "curriculum_module")
    ordering = ("course", "order")
    fieldsets = (
        (None, {"fields": ("course", "mentor", "curriculum_module", "order", "title")}),
        ("Содержание", {"fields": ("description",)}),
        ("Дополнительно", {"fields": ("is_additional", "date", "deadline")}),
        ("Архив", {"fields": ("is_archived", "archived_at")}),
    )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "curriculum_module":
            object_id = request.resolver_match.kwargs.get("object_id")
            if object_id:
                course_id = Lesson.objects.filter(pk=object_id).values_list("course_id", flat=True).first()
                if course_id:
                    kwargs["queryset"] = CurriculumModule.objects.filter(course_id=course_id).order_by("order")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


crm_admin_site.register(LogEntry, CRMLogEntryAdmin)
