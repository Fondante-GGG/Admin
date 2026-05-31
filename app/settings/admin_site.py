from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import date, datetime, timedelta
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from django.conf import settings as django_settings
from django.contrib.auth import password_validation
from django.contrib.admin import AdminSite
from django.http import HttpResponse
from django.http import JsonResponse
from django.http import HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect
from django.urls import path, reverse
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db.models import Count, Prefetch, Q, Sum
from django.db.models.functions import TruncMonth
from django.utils import formats, timezone
from django.utils.dateparse import parse_date
from django.utils.text import slugify
from django.template.response import TemplateResponse
from django.core.paginator import Paginator
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Font

from .models import (
    AccountingAccount,
    AccountingCategory,
    AccountingEntry,
    AccountingProject,
    BillingRecord,
    CalendarEvent,
    CourseContract,
    CourseDrop,
    Cursues,
    Enrollment,
    Lead,
    Mentor,
    Parent,
    Payment,
    Salary,
    Student,
    Task,
    TuitionPayment,
    User,
)
from app.config.models import CRMAbout


def _validation_message(exc: ValidationError) -> str:
    parts = []
    if hasattr(exc, "error_dict"):
        for errs in exc.error_dict.values():
            parts.extend(str(e) for e in errs)
    else:
        parts = list(getattr(exc, "messages", []) or [str(exc)])
    return " ".join(parts)


def _require_valid_password(raw_password: str, user: User) -> str:
    password = (raw_password or "").strip()
    if not password:
        raise ValidationError("Укажите пароль.")
    password_validation.validate_password(password, user)
    return password


@dataclass(frozen=True)
class MonthSeries:
    labels: list[str]
    values: list[float]


def _last_month_starts(months: int) -> list[date]:
    today = timezone.localdate()
    first_of_this_month = today.replace(day=1)
    starts: list[date] = []
    year = first_of_this_month.year
    month = first_of_this_month.month
    for _ in range(months):
        starts.append(date(year, month, 1))
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    return list(reversed(starts))


def _sum_by_month(qs, months: int, field: str) -> MonthSeries:
    month_starts = _last_month_starts(months)
    start = month_starts[0]
    end = (month_starts[-1].replace(day=28) + timedelta(days=4)).replace(day=1)

    grouped = (
        qs.filter(created_at__gte=start, created_at__lt=end)
        .annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(total=Sum(field))
        .order_by("month")
    )
    totals = {row["month"].date(): float(row["total"] or 0) for row in grouped if row["month"]}

    labels: list[str] = []
    values: list[float] = []
    for m in month_starts:
        labels.append(m.strftime("%b %y"))
        values.append(totals.get(m, 0.0))
    return MonthSeries(labels=labels, values=values)


def _count_by_month(qs, months: int) -> MonthSeries:
    month_starts = _last_month_starts(months)
    start = month_starts[0]
    end = (month_starts[-1].replace(day=28) + timedelta(days=4)).replace(day=1)

    grouped = (
        qs.filter(created_at__gte=start, created_at__lt=end)
        .annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(total=Count("id"))
        .order_by("month")
    )
    totals = {row["month"].date(): float(row["total"] or 0) for row in grouped if row["month"]}

    labels: list[str] = []
    values: list[float] = []
    for m in month_starts:
        labels.append(m.strftime("%b"))
        values.append(totals.get(m, 0.0))
    return MonthSeries(labels=labels, values=values)


def _count_by_day(qs, days: int) -> MonthSeries:
    today = timezone.localdate()
    start = today - timedelta(days=days - 1)
    grouped = (
        qs.filter(created_at__date__gte=start, created_at__date__lte=today)
        .values("created_at__date")
        .annotate(total=Count("id"))
        .order_by("created_at__date")
    )
    totals = {row["created_at__date"]: float(row["total"] or 0) for row in grouped}

    labels: list[str] = []
    values: list[float] = []
    for i in range(days):
        d = start + timedelta(days=i)
        labels.append(d.strftime("%d.%m"))
        values.append(totals.get(d, 0.0))
    return MonthSeries(labels=labels, values=values)


def _distinct_paying_students_by_month(qs, months: int) -> MonthSeries:
    """Students with at least one payment in each calendar month (activity proxy)."""
    month_starts = _last_month_starts(months)
    start = month_starts[0]
    end = (month_starts[-1].replace(day=28) + timedelta(days=4)).replace(day=1)

    grouped = (
        qs.filter(created_at__gte=start, created_at__lt=end)
        .annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(total=Count("student", distinct=True))
        .order_by("month")
    )
    totals = {row["month"].date(): float(row["total"] or 0) for row in grouped if row["month"]}

    labels: list[str] = []
    values: list[float] = []
    for m in month_starts:
        labels.append(formats.date_format(datetime.combine(m, datetime.min.time()), "F").title())
        values.append(totals.get(m, 0.0))
    return MonthSeries(labels=labels, values=values)


def _calendar_events_by_day(days: int) -> MonthSeries:
    today = timezone.localdate()
    start = today - timedelta(days=days - 1)
    grouped = (
        CalendarEvent.objects.filter(is_archived=False, start_at__date__gte=start, start_at__date__lte=today)
        .values("start_at__date")
        .annotate(total=Count("id"))
        .order_by("start_at__date")
    )
    totals = {row["start_at__date"]: float(row["total"] or 0) for row in grouped}

    labels: list[str] = []
    values: list[float] = []
    for i in range(days):
        d = start + timedelta(days=i)
        labels.append(d.strftime("%d.%m"))
        values.append(totals.get(d, 0.0))
    return MonthSeries(labels=labels, values=values)


def _compact_money(amount: float) -> str:
    n = float(amount or 0)
    if abs(n) >= 1_000_000:
        return f"{n / 1_000_000:.1f}M c."
    if abs(n) >= 1000:
        return f"{n / 1000:.0f}K c."
    return f"{n:.0f} c."


def _format_int_ru(n: float | int) -> str:
    v = int(round(float(n)))
    if v < 0:
        return "−" + f"{abs(v):,}".replace(",", " ")
    return f"{v:,}".replace(",", " ")


def _format_money_ru(n: float | int) -> str:
    d = Decimal(str(n or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    neg = d < 0
    d = abs(d)
    if d == d.to_integral_value():
        body = f"{int(d):,}".replace(",", " ")
    else:
        t = f"{d:.2f}"
        ip, fp = t.split(".")
        body = f"{int(ip):,}".replace(",", " ") + "," + fp
    return ("−" if neg else "") + body + " с."


class CRMAdminSite(AdminSite):
    site_header = "Codify CRM"
    site_title = "Codify CRM"
    index_title = "Дашборд"
    index_template = "admin/crm_index.html"

    SIDEBAR_SKIP_MODELS: frozenset[str] = frozenset(
        {
            "settings.call",
            "config.crmbilling",
            "config.crmabout",
            "settings.accountingaccount",
            "settings.accountingproject",
            "settings.accountingcategory",
            "settings.task",
            "settings.user",
            "settings.payment",
        }
    )

    SIDEBAR_PRIMARY_ORDER: tuple[str, ...] = (
        "settings.calendarevent",
        "settings.student",
        "settings.mentor",
        "settings.tuitionpayment",
        "settings.debtorenrollment",
        "settings.payment",
        "settings.salary",
        "settings.lead",
        "settings.accountingentry",
    )

    SIDEBAR_SETTINGS_ONLY_MODELS: frozenset[str] = frozenset({"config.crmsetting"})

    SIDEBAR_SITE_SETTINGS_MODELS: frozenset[str] = frozenset(
        {
            "academy.settings",
            "academy.contacts",
            "academy.aboutpage",
            "academy.aboutobjects",
            "academy.aboutobjects2",
            "academy.typecourse",
            "academy.courses",
            "academy.coursespage",
            "academy.coursesmodel",
            "academy.courseapplication",
            "academy.teacher",
            "academy.students",
            "academy.aboutstudents",
            "academy.feedback",
        }
    )

    SIDEBAR_SITE_SETTINGS_ORDER: tuple[str, ...] = (
        "academy.settings",
        "academy.contacts",
        "academy.aboutpage",
        "academy.aboutobjects",
        "academy.aboutobjects2",
        "academy.typecourse",
        "academy.courses",
        "academy.coursespage",
        "academy.coursesmodel",
        "academy.courseapplication",
        "academy.teacher",
        "academy.students",
        "academy.aboutstudents",
        "academy.feedback",
    )

    @staticmethod
    def _role(user) -> str:
        role = getattr(user, "role", "") or ""
        return "Администратор" if role == "Админ" else role

    def _is_manager(self, request) -> bool:
        return self._role(request.user) == "Менеджер"

    def _deny_accounting_for_manager(self, request):
        if self._is_manager(request):
            return HttpResponseForbidden("Managers cannot access accounting.")
        return None

    def has_permission(self, request):
        user = request.user
        if not user.is_authenticated or not user.is_active:
            return False
        return super().has_permission(request)

    def _path_matches(self, request, url: str | None, *, exact: bool = False) -> bool:
        if not url:
            return False
        target = (urlsplit(url).path or "").rstrip("/") or "/"
        current = (request.path or "").rstrip("/") or "/"
        if exact:
            return current == target
        return current == target or current.startswith(f"{target}/")

    def _sidebar_item(
        self,
        request,
        *,
        title: str,
        icon: str,
        url: str | None = None,
        children: list[dict] | None = None,
        exact: bool = False,
        match_url: bool = True,
        force_open: bool = False,
    ) -> dict:
        children = children or []
        child_active = any(child.get("active") or child.get("open") for child in children)
        active = child_active or (match_url and self._path_matches(request, url, exact=exact))
        return {
            "title": title,
            "icon": icon,
            "url": url,
            "children": children,
            "active": active,
            "open": force_open or child_active,
        }

    def _build_sidebar_menu(self, request) -> list[dict]:
        jazzmin_settings = getattr(django_settings, "JAZZMIN_SETTINGS", {})
        icon_map = jazzmin_settings.get("icons", {})
        hidden_models = {m.lower() for m in jazzmin_settings.get("hide_models", [])}
        skip_sidebar = hidden_models | set(self.SIDEBAR_SKIP_MODELS)
        default_child_icon = jazzmin_settings.get("default_icon_children", "far fa-circle")

        model_map: dict[str, dict] = {}
        model_order: list[str] = []
        for app in self.get_app_list(request):
            app_label = app["app_label"].lower()
            for model in app.get("models", []):
                model_str = f"{app_label}.{model['object_name']}".lower()
                if model_str in skip_sidebar:
                    continue
                model_map[model_str] = {
                    "title": model.get("name") or model["object_name"],
                    "icon": icon_map.get(model_str, default_child_icon),
                    "url": model.get("admin_url"),
                }
                model_order.append(model_str)

        index_url = reverse(f"{self.name}:index")
        brand_label = (jazzmin_settings.get("site_brand") or self.site_header or "CRM").strip().upper()

        menu: list[dict] = []
        handled: set[str] = set()
        course_children_keys = frozenset({"settings.groupcourse", "settings.individualcourse", "settings.cursues"})

        def take_model(model_str: str, *, title: str | None = None, icon: str | None = None, url: str | None = None) -> dict | None:
            item = model_map.get(model_str)
            if not item:
                return None
            handled.add(model_str)
            return self._sidebar_item(
                request,
                title=title or item["title"],
                icon=icon or item["icon"],
                url=url or item["url"],
            )

        def append_courses(bucket: list[dict]):
            course_children: list[dict] = []
            group_item = take_model("settings.groupcourse")
            if group_item:
                course_children.append(group_item)
            individual_item = take_model("settings.individualcourse")
            if individual_item:
                course_children.append(individual_item)

            handled.add("settings.cursues")
            if course_children:
                bucket.append(
                    self._sidebar_item(
                        request,
                        title="Курсы",
                        icon=icon_map.get("settings.cursues", "fas fa-book"),
                        children=course_children,
                    )
                )
            else:
                course_item = take_model("settings.cursues", title="Курсы")
                if course_item:
                    bucket.append(course_item)

        menu.append({"heading": True, "title": brand_label})

        org_children = []
        try:
            from app.settings.models import Organization
            orgs = list(Organization.objects.order_by("name").values("id", "name"))
        except Exception:
            orgs = []

        if orgs:
            org_children.append(
                self._sidebar_item(
                    request,
                    title="Все организации",
                    icon="fas fa-angle-right",
                    url=reverse(f"{self.name}:org_switch", args=[0]),
                    match_url=False,
                )
            )
            for org in orgs:
                org_children.append(
                    self._sidebar_item(
                        request,
                        title=org["name"],
                        icon="fas fa-angle-right",
                        url=reverse(f"{self.name}:org_switch", args=[org["id"]]),
                        match_url=False,
                    )
                )
            menu.append(
                self._sidebar_item(
                    request,
                    title="Мои организации",
                    icon="fas fa-map-marker-alt",
                    children=org_children,
                    match_url=False,
                    force_open=True,
                )
            )

        menu.append(
            self._sidebar_item(
                request,
                title="Дашборд",
                icon="fas fa-home",
                url=index_url,
                exact=True,
            )
        )

        # Добавляем ссылку на регистрацию пользователей
        menu.append(
            self._sidebar_item(
                request,
                title="Регистрация пользователей",
                icon="fas fa-user-plus",
                url="/register/",
                match_url=False,
            )
        )

        if "settings.cursues" in model_map:
            append_courses(menu)

        for model_str in self.SIDEBAR_PRIMARY_ORDER:
            if model_str == "settings.calendarevent" and model_str in model_map:
                handled.add(model_str)
                menu.append(
                    self._sidebar_item(
                        request,
                        title=model_map[model_str]["title"],
                        icon=model_map[model_str]["icon"],
                        url=reverse(f"{self.name}:calendar"),
                    )
                )
                continue
            item = take_model(model_str)
            if item:
                menu.append(item)

        for model_str in model_order:
            if (
                model_str in handled
                or model_str in course_children_keys
                or model_str in self.SIDEBAR_SETTINGS_ONLY_MODELS
                or model_str in self.SIDEBAR_SITE_SETTINGS_MODELS
            ):
                continue
            if model_str == "settings.calendarevent":
                handled.add(model_str)
                menu.append(
                    self._sidebar_item(
                        request,
                        title=model_map[model_str]["title"],
                        icon=model_map[model_str]["icon"],
                        url=reverse(f"{self.name}:calendar"),
                    )
                )
                continue
            item = take_model(model_str)
            if item:
                menu.append(item)

        menu.append(
            self._sidebar_item(
                request,
                title="Архив",
                icon="fas fa-archive",
                url=reverse(f"{self.name}:archive_index"),
            )
        )

        site_settings_children: list[dict] = []
        for model_str in self.SIDEBAR_SITE_SETTINGS_ORDER:
            child = take_model(model_str)
            if child:
                site_settings_children.append(child)
        if site_settings_children:
            menu.append(
                self._sidebar_item(
                    request,
                    title="Настройки сайта",
                    icon="fas fa-globe",
                    url=reverse(f"{self.name}:site_settings_index"),
                    children=site_settings_children,
                )
            )

        menu.append({"heading": True, "title": "НАСТРОЙКИ"})

        menu.append(
            self._sidebar_item(
                request,
                title="Настройки",
                icon="fas fa-cog",
                url=reverse(f"{self.name}:settings_index"),
            )
        )

        history_url = reverse(f"{self.name}:admin_logentry_changelist")
        for link in jazzmin_settings.get("custom_links", {}).get("config", []):
            if (link.get("name") or "").strip() == "История действий" and link.get("url"):
                history_url = link["url"]
                break
        menu.append(
            self._sidebar_item(
                request,
                title="История действий",
                icon="fas fa-history",
                url=history_url,
            )
        )

        return menu

    def each_context(self, request):
        context = super().each_context(request)
        context.update(
            {
                "crm_admin_site_name": self.name,
                "crm_sidebar_index_url": reverse(f"{self.name}:index"),
                "crm_sidebar_menu": self._build_sidebar_menu(request),
            }
        )
        return context

    def get_app_list(self, request, app_label=None):
        app_list = super().get_app_list(request, app_label=app_label)
        role = self._role(request.user)

        if role == "Администратор":
            return app_list

        def filter_models(models):
            if role == "Менеджер":
                blocked = {"AccountingEntry", "AccountingAccount", "AccountingProject", "AccountingCategory"}
                return [m for m in models if m.get("object_name") not in blocked]
            if role == "Ментор":
                allowed = {
                    "Cursues",
                    "Student",
                    "Mentor",
                    "StudentPayments",
                    "TuitionPayment",
                    "DebtorEnrollment",
                    "CalendarEvent",
                }
                return [m for m in models if m.get("object_name") in allowed]
            return models

        out = []
        for app in app_list:
            models = filter_models(app.get("models", []))
            if models:
                app2 = dict(app)
                app2["models"] = models
                out.append(app2)
        return out

    @staticmethod
    def _with_query(url: str, **updates) -> str:
        parsed = urlsplit(url)
        params = dict(parse_qsl(parsed.query, keep_blank_values=True))
        for key, value in updates.items():
            if value in (None, ""):
                params.pop(key, None)
            else:
                params[key] = str(value)
        query = urlencode(params)
        return urlunsplit((parsed.scheme, parsed.netloc, parsed.path, query, parsed.fragment))

    def _course_change_url(self, course_id: int, **params) -> str:
        url = reverse(f"{self.name}:settings_groupcourse_change", args=[course_id])
        return self._with_query(url, **params)

    @staticmethod
    def _normalize_phone(value: str) -> str:
        value = (value or "").strip()
        if not value:
            return ""
        return re.sub(r"[^\d+]+", "", value)

    @staticmethod
    def _split_full_name(full_name: str) -> tuple[str, str]:
        parts = [part for part in (full_name or "").strip().split() if part]
        if not parts:
            return "", ""
        if len(parts) == 1:
            return parts[0], ""
        return parts[0], " ".join(parts[1:])

    def _unique_username(self, base: str) -> str:
        base = slugify(base or "") or "student"
        candidate = base[:150]
        suffix = 1
        while User.objects.filter(username=candidate).exists():
            tail = f"-{suffix}"
            candidate = f"{base[:150-len(tail)]}{tail}"
            suffix += 1
        return candidate

    @staticmethod
    def _parse_decimal(value: str, default: Decimal | None = None) -> Decimal | None:
        raw = (value or "").strip().replace(" ", "").replace(",", ".")
        if not raw:
            return default
        try:
            return Decimal(raw)
        except InvalidOperation:
            return default

    def _decode_upload(self, uploaded_file) -> str:
        payload = uploaded_file.read()
        for encoding in ("utf-8-sig", "utf-8", "cp1251"):
            try:
                return payload.decode(encoding)
            except UnicodeDecodeError:
                continue
        return payload.decode("utf-8", errors="ignore")

    def _map_student_upload_rows(self, rows: list[list[str]]):
        aliases = {
            "fio": {"fio", "фио", "full_name", "fullname", "name", "имя", "imya", "student", "student_name"},
            "phone": {"phone", "phone_number", "телефон", "telefon", "номер", "mobile"},
            "tuition": {"tuition", "amount", "sum", "summa", "price", "оплата", "сумма", "tuition_amount"},
            "username": {"username", "логин", "user", "login"},
        }

        def normalize_header(cell: str) -> str:
            return re.sub(r"[^\w]+", "_", (cell or "").strip().lower(), flags=re.UNICODE).strip("_")

        header = [normalize_header(cell) for cell in rows[0]]
        flat_aliases = {item for values in aliases.values() for item in values}
        has_header = any(cell in flat_aliases for cell in header)
        data_rows = rows[1:] if has_header else rows

        mapped_rows = []
        for row in data_rows:
            cells = [str(cell or "").strip() for cell in row]
            if has_header:
                row_map = {}
                for idx, cell in enumerate(cells):
                    if idx >= len(header):
                        continue
                    key = header[idx]
                    row_map[key] = cell

                def pick(alias_key: str) -> str:
                    for alias in aliases[alias_key]:
                        if alias in row_map and row_map[alias]:
                            return row_map[alias]
                    return ""

                mapped_rows.append(
                    {
                        "full_name": pick("fio"),
                        "phone": pick("phone"),
                        "tuition": pick("tuition"),
                        "username": pick("username"),
                    }
                )
            else:
                mapped_rows.append(
                    {
                        "full_name": cells[0] if len(cells) > 0 else "",
                        "phone": cells[1] if len(cells) > 1 else "",
                        "tuition": cells[2] if len(cells) > 2 else "",
                        "username": cells[3] if len(cells) > 3 else "",
                    }
                )
        return mapped_rows

    def _iter_student_upload_rows(self, uploaded_file):
        filename = (getattr(uploaded_file, "name", "") or "").lower()
        if filename.endswith((".xlsx", ".xlsm")):
            try:
                from openpyxl import load_workbook
            except ModuleNotFoundError as exc:
                raise RuntimeError("openpyxl is required for xlsx imports") from exc

            workbook = load_workbook(io.BytesIO(uploaded_file.read()), read_only=True, data_only=True)
            sheet = workbook.active
            rows = [
                ["" if cell is None else str(cell).strip() for cell in row]
                for row in sheet.iter_rows(values_only=True)
                if any(str(cell).strip() for cell in row if cell is not None)
            ]
            return self._map_student_upload_rows(rows)

        text = self._decode_upload(uploaded_file)
        sample = text[:2048]
        delimiter = ","
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
        except csv.Error:
            pass
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = [row for row in reader if any((cell or "").strip() for cell in row)]
        return self._map_student_upload_rows(rows)

    @staticmethod
    def _contract_periods(course: Cursues) -> str:
        if not course.start:
            return "Период не указан"
        if course.duration_days:
            end_date = course.start + timedelta(days=max(int(course.duration_days) - 1, 0))
            return f"с {course.start.strftime('%d %b %Y')} по {end_date.strftime('%d %b %Y')}"
        return f"с {course.start.strftime('%d %b %Y')}"

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "settings/course/<int:course_id>/students.csv",
                self.admin_view(self.course_students_csv),
                name="course_students_csv",
            ),
            path(
                "settings/course/create/",
                self.admin_view(self.course_quick_create),
                name="course_quick_create",
            ),
            path(
                "settings/course/<int:course_id>/update/",
                self.admin_view(self.course_update),
                name="course_update",
            ),
            path(
                "settings/course/<int:course_id>/students/upload/",
                self.admin_view(self.course_students_upload),
                name="course_students_upload",
            ),
            path(
                "settings/course/<int:course_id>/status/",
                self.admin_view(self.course_status_update),
                name="course_status_update",
            ),
            path(
                "settings/course/<int:course_id>/contracts/generate/",
                self.admin_view(self.course_contracts_generate),
                name="course_contracts_generate",
            ),
            path("archive/", self.admin_view(self.archive_index), name="archive_index"),
            path("calendar/", self.admin_view(self.calendar_view), name="calendar"),
            path("calendar/events/", self.admin_view(self.calendar_events), name="calendar_events"),
            path("calendar/events/create/", self.admin_view(self.calendar_event_create), name="calendar_event_create"),
            path("students.csv", self.admin_view(self.students_csv), name="students_csv"),
            path("students.pdf", self.admin_view(self.students_pdf), name="students_pdf"),
            path("students/<int:student_id>/drawer/", self.admin_view(self.student_drawer), name="student_drawer"),
            path("mentors/export.xlsx", self.admin_view(self.mentors_export_xlsx), name="mentors_export_xlsx"),
            path("mentors/salary/", self.admin_view(self.mentors_salary), name="mentors_salary"),
            path("salary/create/", self.admin_view(self.salary_create), name="salary_create"),
            path("salary/<int:salary_id>/delete/", self.admin_view(self.salary_delete), name="salary_delete"),
            path("mentors/<int:mentor_id>/drawer/", self.admin_view(self.mentor_drawer), name="mentor_drawer"),
            path("mentors/create/", self.admin_view(self.mentor_quick_create), name="mentor_quick_create"),
            path("students/create/", self.admin_view(self.student_quick_create), name="student_quick_create"),
            path("parents/create/", self.admin_view(self.parent_quick_create), name="parent_quick_create"),
            path("about/", self.admin_view(self.about_view), name="about"),
            path("settings/", self.admin_view(self.settings_index), name="settings_index"),
            path(
                "site-settings/",
                self.admin_view(self.site_settings_index),
                name="site_settings_index",
            ),
            path("accounting/meta/", self.admin_view(self.accounting_meta), name="accounting_meta"),
            path("accounting/entry/create/", self.admin_view(self.accounting_entry_create), name="accounting_entry_create"),
            path("accounting/transfer/create/", self.admin_view(self.accounting_transfer_create), name="accounting_transfer_create"),
            path("org/switch/<int:org_id>/", self.admin_view(self.org_switch), name="org_switch"),
        ]
        return custom + urls

    def settings_index(self, request):
        role = self._role(request.user)
        can_edit = role in ("Администратор", "Менеджер") and request.user.is_staff

        def _link(name, url):
            return {"name": name, "url": url}

        def _admin_url(model_name, action="changelist"):
            try:
                return reverse(f"{self.name}:settings_{model_name}_{action}")
            except Exception:
                return None

        def _config_url(model_name, action="changelist"):
            try:
                return reverse(f"{self.name}:config_{model_name}_{action}")
            except Exception:
                return None

        sections = [
            {
                "title": "Администрирование системы",
                "icon": "fas fa-shield-alt",
                "color": "#8b5cf6",
                "items": [
                    _link("Организации", _admin_url("organization")),
                    _link("Персонал", _admin_url("user")),
                    _link("Настройки организации", _config_url("crmsetting")),
                    _link("Домены", None),
                ],
            },
            {
                "title": "Система обучения",
                "icon": "fas fa-graduation-cap",
                "color": "#f59e0b",
                "items": [
                    _link("Предметы/направления", _admin_url("cursues")),
                    _link("Электронная библиотека", None),
                ],
            },
            {
                "title": "Геймификация",
                "icon": "fas fa-gem",
                "color": "#ec4899",
                "items": [
                    _link("Валюты", None),
                ],
            },
            {
                "title": "Другие настройки",
                "icon": "fas fa-tools",
                "color": "#6b7280",
                "items": [
                    _link("Настройки формы регистрации", None),
                    _link("Шаблоны документов", None),
                    _link("Дополнительные поля для клиентов", None),
                ],
            },
            {
                "title": "Лид менеджмент",
                "icon": "fas fa-briefcase",
                "color": "#10b981",
                "items": [
                    _link("Воронки продаж", None),
                    _link("Кошельки", None),
                    _link("Причины потери лида", None),
                    _link("Специальные поля для лидов", None),
                    _link("Дополнительные поля для лидов", None),
                    _link("Интеграция с другими сайтами", None),
                    _link("Метки/Тэги", None),
                ],
            },
            {
                "title": "Платежи",
                "icon": "fas fa-dollar-sign",
                "color": "#059669",
                "items": [
                    _link("Способы оплаты", _admin_url("payment")),
                ],
            },
            {
                "title": "Филиальность",
                "icon": "fas fa-map-marker-alt",
                "color": "#3b82f6",
                "items": [
                    _link("Организации", _admin_url("organization")),
                ],
            },
            {
                "title": "Интеграции",
                "icon": "fas fa-puzzle-piece",
                "color": "#6366f1",
                "items": [
                    _link("Мессенджеры", None),
                    _link("АТС", _admin_url("call")),
                ],
            },
            {
                "title": "Интерфейс",
                "icon": "fas fa-desktop",
                "color": "#8b5cf6",
                "items": [
                    _link("Боковое меню", None),
                ],
            },
            {
                "title": "Уведомления",
                "icon": "fas fa-bell",
                "color": "#ef4444",
                "items": [
                    _link("Телеграм-бот", None),
                ],
            },
        ]

        for section in sections:
            section["items"] = [i for i in section["items"] if i["url"] or not can_edit or True]

        context = {
            "title": "Настройки",
            "sections": sections,
            "can_edit": can_edit,
            "storage_used": "0.0",
            "storage_total": "5 GB",
            "accounts_used": 1,
            "accounts_total": 200,
            "accounts_percent": 0,
            **self.each_context(request),
        }
        return TemplateResponse(request, "admin/settings_index.html", context)

    def site_settings_index(self, request):
        def _item(name, model_name):
            try:
                return {
                    "name": name,
                    "list_url": reverse(f"{self.name}:academy_{model_name}_changelist"),
                    "add_url": reverse(f"{self.name}:academy_{model_name}_add"),
                }
            except Exception:
                return {"name": name, "list_url": None, "add_url": None}

        sections = [
            {
                "title": "Главная страница",
                "icon": "fas fa-home",
                "color": "#3b82f6",
                "items": [_item("Баннер и блоки", "settings")],
            },
            {
                "title": "Контакты",
                "icon": "fas fa-address-book",
                "color": "#10b981",
                "items": [_item("Контакты и адреса", "contacts")],
            },
            {
                "title": "Страница «О нас»",
                "icon": "fas fa-info-circle",
                "color": "#8b5cf6",
                "items": [
                    _item("Баннер и текст", "aboutpage"),
                    _item("Данные 1", "aboutobjects"),
                    _item("Данные 2", "aboutobjects2"),
                ],
            },
            {
                "title": "Курсы",
                "icon": "fas fa-book",
                "color": "#f59e0b",
                "items": [
                    _item("Типы курсов", "typecourse"),
                    _item("Курсы", "courses"),
                    _item("Страница курсов", "coursespage"),
                    _item("Модальные окна", "coursesmodel"),
                    _item("Заявки на курсы", "courseapplication"),
                ],
            },
            {
                "title": "Преподаватели",
                "icon": "fas fa-chalkboard-teacher",
                "color": "#ec4899",
                "items": [_item("Преподаватели", "teacher")],
            },
            {
                "title": "Студенты",
                "icon": "fas fa-user-graduate",
                "color": "#6366f1",
                "items": [
                    _item("Страница студентов", "students"),
                    _item("Выпускники", "aboutstudents"),
                ],
            },
            {
                "title": "Обратная связь",
                "icon": "fas fa-envelope",
                "color": "#14b8a6",
                "items": [_item("Сообщения с сайта", "feedback")],
            },
        ]

        context = {
            "title": "Настройки сайта",
            "sections": sections,
            "can_edit": request.user.is_staff,
            **self.each_context(request),
        }
        return TemplateResponse(request, "admin/site_settings_index.html", context)

    def org_switch(self, request, org_id: int):
        from app.settings.models import Organization
        if org_id:
            org = get_object_or_404(Organization, pk=org_id)
            request.session["current_org_id"] = org.pk
            request.session["current_org_name"] = org.name
            messages.success(request, f"Организация изменена на: {org.name}")
        else:
            request.session.pop("current_org_id", None)
            request.session.pop("current_org_name", None)
            messages.success(request, "Выбраны все организации")
        referer = request.META.get("HTTP_REFERER") or reverse(f"{self.name}:index")
        return redirect(referer)

    def about_view(self, request):
        about = CRMAbout.objects.order_by("-updated_at").first()
        role = self._role(request.user)
        can_edit = role in ("Администратор", "Менеджер") and request.user.is_staff
        context = {
            "title": "О нас",
            "about": about,
            "can_edit": can_edit,
            **self.each_context(request),
        }
        return TemplateResponse(request, "admin/about.html", context)

    def course_quick_create(self, request):
        if request.method != "POST":
            return JsonResponse({"ok": False, "error": "method not allowed"}, status=405)
        if self._role(request.user) not in {"Администратор", "Менеджер", "Ментор"}:
            return HttpResponseForbidden("Недостаточно прав.")

        def safe_next(default_url: str) -> str:
            next_url = (request.POST.get("next") or "").strip()
            if next_url.startswith("/") and not next_url.startswith("//"):
                return next_url
            return default_url

        def parse_positive_int(value: str, default: int, minimum: int = 0) -> int:
            try:
                return max(minimum, int((value or "").strip()))
            except (TypeError, ValueError):
                return default

        course_type = (request.POST.get("course_type") or Cursues.CourseType.GROUP).strip()
        if course_type not in {Cursues.CourseType.GROUP, Cursues.CourseType.INDIVIDUAL}:
            course_type = Cursues.CourseType.GROUP

        list_url_name = (
            "settings_individualcourse_changelist"
            if course_type == Cursues.CourseType.INDIVIDUAL
            else "settings_groupcourse_changelist"
        )
        next_url = safe_next(reverse(f"{self.name}:{list_url_name}"))

        title = (request.POST.get("title") or "").strip()
        if not title:
            messages.error(request, "Укажите название курса.")
            return redirect(next_url)

        allowed_statuses = {
            "Подготовка",
            "Готов к запуску",
            "Запущен",
            "Приостановлен",
            "Закончен",
            "Архивирован",
        }
        status = (request.POST.get("status") or "Подготовка").strip()
        if status not in allowed_statuses:
            status = "Подготовка"

        start = parse_date((request.POST.get("start") or "").strip()) or timezone.localdate()
        duration_months = parse_positive_int(request.POST.get("duration_months") or "", default=6, minimum=1)
        capacity = parse_positive_int(request.POST.get("capacity") or "", default=10, minimum=0)
        org_id = request.session.get("current_org_id")

        course = Cursues.objects.create(
            organization_id=org_id or None,
            title=title,
            course_type=course_type,
            subject=(request.POST.get("subject") or "").strip(),
            status=status,
            start=start,
            duration_days=duration_months * 30,
            capacity=capacity,
            room=(request.POST.get("room") or "").strip(),
            schedule_note="",
            is_archived=status == "Архивирован",
            archived_at=timezone.now() if status == "Архивирован" else None,
        )

        messages.success(request, f"Курс «{course.title}» добавлен.")
        return redirect(next_url)

    def accounting_meta(self, request):
        denied = self._deny_accounting_for_manager(request)
        if denied:
            return denied
        accounts = list(AccountingAccount.objects.order_by("title").values("id", "title"))
        projects = list(AccountingProject.objects.order_by("title").values("id", "title"))
        categories = list(AccountingCategory.objects.order_by("title").values("id", "title"))
        return JsonResponse({"accounts": accounts, "projects": projects, "categories": categories})

    def accounting_entry_create(self, request):
        denied = self._deny_accounting_for_manager(request)
        if denied:
            return denied
        if request.method != "POST":
            return JsonResponse({"ok": False, "error": "method not allowed"}, status=405)
        entry_type = (request.POST.get("entry_type") or "").strip()
        if entry_type not in (AccountingEntry.Type.INCOME, AccountingEntry.Type.EXPENSE):
            return JsonResponse({"ok": False, "error": "invalid entry_type"}, status=400)

        account_id = (request.POST.get("account") or "").strip()
        amount_raw = (request.POST.get("amount") or "").strip()
        operated_at_raw = (request.POST.get("operated_at") or "").strip()
        project_id = (request.POST.get("project") or "").strip()
        category_id = (request.POST.get("category") or "").strip()
        title = (request.POST.get("title") or "").strip()

        if not account_id:
            return JsonResponse({"ok": False, "error": "account required"}, status=400)
        try:
            amount = float(amount_raw or 0)
        except ValueError:
            return JsonResponse({"ok": False, "error": "invalid amount"}, status=400)
        if amount <= 0:
            return JsonResponse({"ok": False, "error": "amount must be > 0"}, status=400)

        operated_at = timezone.now()
        if operated_at_raw:
            dt = timezone.datetime.fromisoformat(operated_at_raw)
            operated_at = dt if timezone.is_aware(dt) else timezone.make_aware(dt, timezone.get_current_timezone())

        AccountingEntry.objects.create(
            entry_type=entry_type,
            title=title or ("Приход" if entry_type == AccountingEntry.Type.INCOME else "Расход"),
            amount=amount,
            operated_at=operated_at,
            account_id=int(account_id),
            project_id=int(project_id) if project_id else None,
            category_id=int(category_id) if category_id else None,
        )
        return JsonResponse({"ok": True})

    def course_update(self, request, course_id: int):
        if request.method != "POST":
            return JsonResponse({"ok": False, "error": "method not allowed"}, status=405)
        course = get_object_or_404(Cursues, pk=course_id)

        title = (request.POST.get("title") or "").strip()
        subject = (request.POST.get("subject") or "").strip()
        room = (request.POST.get("room") or "").strip()
        status = (request.POST.get("status") or "").strip()
        schedule_note = (request.POST.get("schedule_note") or "").strip()

        if title:
            course.title = title

        capacity_raw = (request.POST.get("capacity") or "").strip()
        if capacity_raw:
            try:
                course.capacity = max(0, int(capacity_raw))
            except ValueError:
                pass

        start_raw = (request.POST.get("start") or "").strip()
        if start_raw:
            try:
                parsed = (
                    timezone.datetime.fromisoformat(start_raw).date()
                    if "T" in start_raw
                    else date.fromisoformat(start_raw)
                )
            except ValueError:
                parsed = None
            if parsed:
                course.start = parsed

        duration_raw = (request.POST.get("duration_months") or "").strip()
        if duration_raw:
            try:
                duration_months = max(0, int(duration_raw))
                course.duration_days = duration_months * 30
            except ValueError:
                pass

        price_raw = (request.POST.get("price") or "").strip()
        if price_raw:
            try:
                course.price = Decimal(price_raw)
            except (InvalidOperation, ValueError):
                pass

        if status:
            course.status = status
            if status == "Архивирован":
                course.is_archived = True
                course.archived_at = course.archived_at or timezone.now()
            elif course.status != "Архивирован":
                course.is_archived = False
                course.archived_at = None
        course.subject = subject
        course.room = room
        course.schedule_note = schedule_note
        course.save()

        next_url = (request.POST.get("next") or "").strip() or reverse(f"{self.name}:settings_groupcourse_change", args=[course.pk])
        return redirect(next_url)

    def course_status_update(self, request, course_id: int):
        if request.method != "POST":
            return JsonResponse({"ok": False, "error": "method not allowed"}, status=405)
        course = get_object_or_404(Cursues, pk=course_id)
        action = (request.POST.get("action") or "").strip()

        if action == "finish":
            course.status = "Закончен"
            course.save(update_fields=["status"])
        elif action == "archive":
            course.status = "Архивирован"
            course.is_archived = True
            course.archived_at = timezone.now()
            course.save(update_fields=["status", "is_archived", "archived_at"])

        if action == "archive":
            next_url = reverse(f"{self.name}:settings_groupcourse_changelist")
            next_url = f"{next_url}?archived=1"
        else:
            next_url = (request.POST.get("next") or "").strip() or reverse(
                f"{self.name}:settings_groupcourse_change",
                args=[course.pk],
            )
        return redirect(next_url)

    def course_students_upload(self, request, course_id: int):
        if request.method != "POST":
            return JsonResponse({"ok": False, "error": "method not allowed"}, status=405)

        course = get_object_or_404(Cursues, pk=course_id)
        upload = request.FILES.get("students_file")
        if not upload:
            return redirect(self._course_change_url(course.pk, panel="students", toast="Файл не выбран"))

        try:
            rows = self._iter_student_upload_rows(upload)
        except Exception:
            return redirect(self._course_change_url(course.pk, panel="students", toast="Не удалось прочитать файл"))

        if not rows:
            return redirect(self._course_change_url(course.pk, panel="students", toast="Файл пустой"))

        created_students = 0
        created_enrollments = 0
        skipped = 0
        updated = 0

        with transaction.atomic():
            for row in rows:
                full_name = (row.get("full_name") or "").strip()
                phone = self._normalize_phone(row.get("phone") or "")
                username_hint = (row.get("username") or "").strip()
                tuition_amount = self._parse_decimal(row.get("tuition") or "", default=course.price) or course.price

                if not full_name and not phone and not username_hint:
                    skipped += 1
                    continue

                first_name, last_name = self._split_full_name(full_name)

                user = None
                if phone:
                    user = User.objects.filter(phone_number=phone).first()
                if user is None and username_hint:
                    user = User.objects.filter(username=username_hint).first()

                created_user = False
                if user is None:
                    username_base = username_hint or phone or full_name or f"student-{timezone.now().timestamp()}"
                    user = User.objects.create(
                        username=self._unique_username(username_base),
                        first_name=first_name,
                        last_name=last_name,
                        phone_number=phone or username_hint or "0000000000",
                        role="Студент",
                        is_active=True,
                        is_staff=False,
                    )
                    user.set_unusable_password()
                    user.save(update_fields=["password"])
                    created_user = True
                else:
                    dirty_fields = []
                    if first_name and not user.first_name:
                        user.first_name = first_name
                        dirty_fields.append("first_name")
                    if last_name and not user.last_name:
                        user.last_name = last_name
                        dirty_fields.append("last_name")
                    if phone and user.phone_number != phone:
                        user.phone_number = phone
                        dirty_fields.append("phone_number")
                    if dirty_fields:
                        user.save(update_fields=dirty_fields)

                student, student_created = Student.objects.get_or_create(user=user)
                if student_created:
                    created_students += 1

                enrollment, enrollment_created = Enrollment.objects.get_or_create(
                    student=student,
                    course=course,
                    defaults={"tuition_amount": tuition_amount},
                )
                if not enrollment_created and tuition_amount is not None and enrollment.tuition_amount != tuition_amount:
                    enrollment.tuition_amount = tuition_amount
                    enrollment.save(update_fields=["tuition_amount"])
                    updated += 1

                course.students.add(student)
                if enrollment_created:
                    created_enrollments += 1

        toast = (
            f"Импорт завершен: студентов {created_students}, "
            f"зачислений {created_enrollments}, обновлено {updated}, пропущено {skipped}"
        )
        return redirect(self._course_change_url(course.pk, panel="students", toast=toast))

    def course_contracts_generate(self, request, course_id: int):
        if request.method != "POST":
            return JsonResponse({"ok": False, "error": "method not allowed"}, status=405)

        course = get_object_or_404(Cursues, pk=course_id)
        enrollments = (
            Enrollment.objects.filter(course=course)
            .select_related("student__user")
            .order_by("student__user__first_name", "student__user__username")
        )

        created = 0
        updated = 0
        periods = self._contract_periods(course)

        with transaction.atomic():
            for enrollment in enrollments:
                student_name = enrollment.student.user.get_full_name() or enrollment.student.user.username
                paid_total = enrollment.paid_total or Decimal("0")
                debt_total = enrollment.debt or Decimal("0")
                document_text = (
                    f"Контракт на обучение\n"
                    f"Курс: {course.title}\n"
                    f"Студент: {student_name}\n"
                    f"Период: {periods}\n"
                    f"Сумма: {enrollment.tuition_amount} с.\n"
                    f"Оплачено: {paid_total} с.\n"
                    f"Долг: {debt_total} с."
                )
                contract, contract_created = CourseContract.objects.update_or_create(
                    course=course,
                    student=enrollment.student,
                    defaults={
                        "periods": periods,
                        "amount_snapshot": enrollment.tuition_amount,
                        "paid_snapshot": paid_total,
                        "debt_snapshot": debt_total,
                        "document_text": document_text,
                    },
                )
                if contract_created:
                    created += 1
                else:
                    updated += 1

        toast = f"Контракты готовы: создано {created}, обновлено {updated}"
        return redirect(self._course_change_url(course.pk, panel="payments", payment_tab="contracts", toast=toast))

    def accounting_transfer_create(self, request):
        denied = self._deny_accounting_for_manager(request)
        if denied:
            return denied
        if request.method != "POST":
            return JsonResponse({"ok": False, "error": "method not allowed"}, status=405)
        from_account = (request.POST.get("from_account") or "").strip()
        to_account = (request.POST.get("to_account") or "").strip()
        amount_raw = (request.POST.get("amount") or "").strip()
        operated_at_raw = (request.POST.get("operated_at") or "").strip()
        project_id = (request.POST.get("project") or "").strip()
        category_id = (request.POST.get("category") or "").strip()
        title = (request.POST.get("title") or "").strip() or "Перевод"

        if not from_account or not to_account:
            return JsonResponse({"ok": False, "error": "accounts required"}, status=400)
        if from_account == to_account:
            return JsonResponse({"ok": False, "error": "accounts must differ"}, status=400)
        try:
            amount = float(amount_raw or 0)
        except ValueError:
            return JsonResponse({"ok": False, "error": "invalid amount"}, status=400)
        if amount <= 0:
            return JsonResponse({"ok": False, "error": "amount must be > 0"}, status=400)

        operated_at = timezone.now()
        if operated_at_raw:
            dt = timezone.datetime.fromisoformat(operated_at_raw)
            operated_at = dt if timezone.is_aware(dt) else timezone.make_aware(dt, timezone.get_current_timezone())

        group = timezone.now().strftime("tr%Y%m%d%H%M%S%f")
        with transaction.atomic():
            AccountingEntry.objects.create(
                entry_type=AccountingEntry.Type.EXPENSE,
                title=title,
                amount=amount,
                operated_at=operated_at,
                account_id=int(from_account),
                project_id=int(project_id) if project_id else None,
                category_id=int(category_id) if category_id else None,
                transfer_group=group,
            )
            AccountingEntry.objects.create(
                entry_type=AccountingEntry.Type.INCOME,
                title=title,
                amount=amount,
                operated_at=operated_at,
                account_id=int(to_account),
                project_id=int(project_id) if project_id else None,
                category_id=int(category_id) if category_id else None,
                transfer_group=group,
            )
        return JsonResponse({"ok": True})

    def course_students_csv(self, request, course_id: int):
        course = Cursues.objects.get(pk=course_id)
        enrollments = (
            Enrollment.objects.select_related("student__user")
            .filter(course=course)
            .order_by("student__user__username")
        )

        lines = ["ФИО,Телефон,Курс"]
        for e in enrollments:
            user = e.student.user
            full_name = (user.get_full_name() or user.username).replace(",", " ")
            phone = (user.phone_number or "").replace(",", " ")
            lines.append(f"{full_name},{phone},{course.title.replace(',', ' ')}")

        content = "\n".join(lines)
        resp = HttpResponse(content, content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="course_{course_id}_students.csv"'
        return resp

    def archive_index(self, request):
        def archive_url(model_name: str) -> str:
            return f"{reverse(f'{self.name}:settings_{model_name}_changelist')}?archived=1"

        sections = [
            {
                "title": "Люди",
                "items": [
                    {"title": "Студенты", "url": archive_url("student")},
                    {"title": "Менторы", "url": archive_url("mentor")},
                    {"title": "Лиды", "url": archive_url("lead")},
                ],
            },
            {
                "title": "Курсы и занятия",
                "items": [
                    {"title": "Курсы", "url": archive_url("cursues")},
                    {"title": "Календарь", "url": archive_url("calendarevent")},
                ],
            },
            {
                "title": "Коммуникации",
                "items": [
                    {"title": "Звонки", "url": archive_url("call")},
                    {"title": "Задачи", "url": archive_url("task")},
                ],
            },
        ]
        if not self._is_manager(request):
            sections.append(
                {
                    "title": "Финансы",
                    "items": [
                        {"title": "Бухгалтерия", "url": archive_url("accountingentry")},
                    ],
                }
            )
        context = {"title": "Архив", "sections": sections, **self.each_context(request)}
        return TemplateResponse(request, "admin/archive_index.html", context)

    def calendar_view(self, request):
        courses = Cursues.objects.filter(is_archived=False).order_by("title")
        mentors = Mentor.objects.select_related("user").order_by("user__username")
        subjects = (
            Cursues.objects.exclude(subject="")
            .values_list("subject", flat=True)
            .distinct()
            .order_by("subject")
        )
        context = {
            "title": "Календарь активных курсов",
            "courses": courses,
            "mentors": mentors,
            "subjects": subjects,
            **self.each_context(request),
        }
        return TemplateResponse(request, "admin/calendar.html", context)

    def calendar_events(self, request):
        qs = CalendarEvent.objects.filter(is_archived=False)

        course = request.GET.get("course")
        if course:
            # allow both id and title
            if course.isdigit():
                qs = qs.filter(course_id=course)
            else:
                qs = qs.filter(course__title=course)

        subject = request.GET.get("subject")
        if subject:
            qs = qs.filter(course__subject=subject)

        mentor_id = request.GET.get("mentor")
        if mentor_id:
            if mentor_id.isdigit():
                qs = qs.filter(course__mentors__id=mentor_id).distinct()
            else:
                qs = qs.filter(course__mentors__user__username=mentor_id).distinct()

        events = []
        for e in qs.select_related("course").order_by("start_at")[:2000]:
            color = None
            if e.course and e.course.subject:
                palette = {
                    "English": "#9ee7e5",
                    "Math": "#b7c3d9",
                    "Python": "#a7f3d0",
                    "Design": "#f5b48f",
                }
                color = palette.get(e.course.subject)

            events.append(
                {
                    "id": e.id,
                    "title": e.title,
                    "start": e.start_at.isoformat(),
                    "end": e.end_at.isoformat() if e.end_at else None,
                    "backgroundColor": color,
                    "borderColor": color,
                }
            )
        return JsonResponse(events, safe=False)

    def calendar_event_create(self, request):
        if request.method != "POST":
            return JsonResponse({"ok": False, "error": "method not allowed"}, status=405)
        title = (request.POST.get("title") or "").strip()
        start = (request.POST.get("start_at") or "").strip()
        end = (request.POST.get("end_at") or "").strip()
        course_id = (request.POST.get("course") or "").strip() or None
        location = (request.POST.get("location") or "").strip()
        online_link = (request.POST.get("online_link") or "").strip()
        description = (request.POST.get("description") or "").strip()

        if not title or not start:
            return JsonResponse({"ok": False, "error": "title/start_at required"}, status=400)

        try:
            start_dt = timezone.datetime.fromisoformat(start)
            if timezone.is_naive(start_dt):
                start_dt = timezone.make_aware(start_dt, timezone.get_current_timezone())
        except ValueError:
            return JsonResponse({"ok": False, "error": "invalid start_at"}, status=400)

        end_dt = None
        if end:
            try:
                end_dt = timezone.datetime.fromisoformat(end)
                if timezone.is_naive(end_dt):
                    end_dt = timezone.make_aware(end_dt, timezone.get_current_timezone())
            except ValueError:
                return JsonResponse({"ok": False, "error": "invalid end_at"}, status=400)

        ev = CalendarEvent.objects.create(
            title=title,
            course_id=course_id,
            start_at=start_dt,
            end_at=end_dt,
            location=location,
            online_link=online_link,
            description=description,
        )
        return JsonResponse({"ok": True, "id": ev.id})

    def _mentor_changelist_qs(self, request):
        qs = Mentor.objects.select_related("user").order_by("-id")
        q = (request.GET.get("q") or "").strip()
        if q:
            text_q = (
                Q(user__first_name__icontains=q)
                | Q(user__last_name__icontains=q)
                | Q(user__username__icontains=q)
                | Q(user__phone_number__icontains=q)
                | Q(user__email__icontains=q)
            )
            if q.isdigit():
                qs = qs.filter(Q(pk=int(q)) | text_q)
            else:
                qs = qs.filter(text_q)
        return qs.distinct()

    def _can_manage_mentors(self, request) -> bool:
        if not request.user.is_authenticated or not request.user.is_staff:
            return False
        return self._role(request.user) in ("Администратор", "Менеджер")

    def mentor_drawer(self, request, mentor_id: int):
        if not self._can_manage_mentors(request):
            return HttpResponseForbidden()
        mentor = (
            Mentor.objects.select_related("user")
            .prefetch_related("cursues_set")
            .get(pk=mentor_id)
        )
        user = mentor.user
        courses = list(mentor.cursues_set.filter(is_archived=False).order_by("-created_at")[:20])
        context = {
            "mentor": mentor,
            "user": user,
            "courses": courses,
            "title": "Информация о менторе",
            **self.each_context(request),
        }
        return TemplateResponse(request, "admin/mentor_drawer.html", context)

    def mentors_export_xlsx(self, request):
        if not self._can_manage_mentors(request):
            return HttpResponseForbidden()
        qs = self._mentor_changelist_qs(request)[:5000]
        wb = Workbook()
        ws = wb.active
        ws.title = "Менторы"
        headers = ("ID", "ФИО", "Курс", "Статус", "Аккаунт", "Телефон", "Почта")
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        course_qs = Cursues.objects.filter(is_archived=False).only("title")
        for m in qs.prefetch_related(Prefetch("cursues_set", queryset=course_qs)):
            u = m.user
            names = [c.title for c in m.cursues_set.all()[:12]]
            course_str = ", ".join(names) if names else "—"
            status_lbl = "Активный" if u.is_active else "Неактивен"
            account = "Да" if u.is_active else "Нет"
            ws.append(
                [
                    m.pk,
                    u.get_full_name() or u.username,
                    course_str,
                    status_lbl,
                    account,
                    u.phone_number or "",
                    u.email or "",
                ]
            )
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="mentors-export.xlsx"'
        return resp

    def mentors_salary(self, request):
        if not self._can_manage_mentors(request):
            return HttpResponseForbidden()

        # Export to Excel (old calculator)
        if request.GET.get("export") == "1":
            return self._mentors_salary_export(request)

        changelist_url = reverse(f"{self.name}:settings_mentor_changelist")
        today = timezone.localdate()
        start_default = today - timedelta(days=30)

        date_from = (request.GET.get("date_from") or "").strip()
        date_to = (request.GET.get("date_to") or "").strip()
        mentor_id = (request.GET.get("mentor") or "").strip()
        course_id = (request.GET.get("course") or "").strip()
        q = (request.GET.get("q") or "").strip()

        qs = Salary.objects.select_related("mentor__user", "course").order_by("-created_at")

        if date_from:
            try:
                d0 = date.fromisoformat(date_from)
                qs = qs.filter(created_at__date__gte=d0)
            except ValueError:
                pass
        if date_to:
            try:
                d1 = date.fromisoformat(date_to)
                qs = qs.filter(created_at__date__lte=d1)
            except ValueError:
                pass
        if mentor_id and mentor_id.isdigit():
            qs = qs.filter(mentor_id=int(mentor_id))
        if course_id and course_id.isdigit():
            qs = qs.filter(course_id=int(course_id))
        if q:
            qs = qs.filter(
                Q(mentor__user__first_name__icontains=q)
                | Q(mentor__user__last_name__icontains=q)
                | Q(mentor__user__username__icontains=q)
                | Q(comment__icontains=q)
            )

        total_amount = float(qs.aggregate(total=Sum("amount"))["total"] or 0)

        mentors_qs = Mentor.objects.select_related("user").filter(user__is_active=True).order_by("user__last_name", "user__first_name")
        courses_qs = Cursues.objects.filter(is_archived=False).order_by("title")

        page_size = 20
        paginator = Paginator(qs, page_size)
        page_number = request.GET.get("p") or 1
        page_obj = paginator.get_page(page_number)

        ctx = {
            "title": "Зарплата",
            "changelist_url": changelist_url,
            "salary_page": page_obj,
            "total_amount": total_amount,
            "mentors": mentors_qs,
            "courses": courses_qs,
            "filters": {
                "date_from": date_from or start_default.isoformat(),
                "date_to": date_to or today.isoformat(),
                "mentor": mentor_id,
                "course": course_id,
                "q": q,
            },
            "tab": request.GET.get("tab") or "payments",
            **self.each_context(request),
        }
        return TemplateResponse(request, "admin/mentor_salary.html", ctx)

    def _mentors_salary_export(self, request):
        """Export salary data to Excel (old calculator logic preserved)."""
        raw_from = (request.GET.get("date_from") or "").strip()
        raw_to = (request.GET.get("date_to") or "").strip()
        try:
            d0 = date.fromisoformat(raw_from) if raw_from else timezone.localdate() - timedelta(days=30)
            d1 = date.fromisoformat(raw_to) if raw_to else timezone.localdate()
        except ValueError:
            d0 = timezone.localdate() - timedelta(days=30)
            d1 = timezone.localdate()
        if d0 > d1:
            d0, d1 = d1, d0
        events = (
            CalendarEvent.objects.filter(
                course__isnull=False,
                course__is_archived=False,
                is_archived=False,
                start_at__date__gte=d0,
                start_at__date__lte=d1,
            )
            .select_related("course")
            .prefetch_related("course__mentors__user")
            .order_by("start_at")
        )
        summary: dict[int, dict] = {}
        rows_detail: list[dict] = []
        for ev in events:
            course = ev.course
            if not course:
                continue
            for mentor in course.mentors.all():
                u = mentor.user
                name = u.get_full_name() or u.username
                mid = mentor.pk
                if mid not in summary:
                    summary[mid] = {"name": name, "count": 0}
                summary[mid]["count"] += 1
                end_s = ""
                if ev.end_at:
                    end_s = timezone.localtime(ev.end_at).strftime("%d.%m.%Y %H:%M")
                rows_detail.append(
                    {
                        "mentor": name,
                        "mentor_id": mid,
                        "course": course.title,
                        "title": ev.title,
                        "start": timezone.localtime(ev.start_at).strftime("%d.%m.%Y %H:%M"),
                        "end": end_s,
                    }
                )
        wb = Workbook()
        ws0 = wb.active
        ws0.title = "Сводка"
        ws0.append(("ID ментора", "ФИО", "Количество занятий", "Примечание"))
        for cell in ws0[1]:
            cell.font = Font(bold=True)
        note = "В CRM не заданы ставки за занятие — итоговую сумму начисляйте вручную по внутренним правилам."
        for mid, data in sorted(summary.items(), key=lambda x: x[1]["name"].lower()):
            ws0.append((mid, data["name"], data["count"], note))
        ws1 = wb.create_sheet("Занятия")
        ws1.append(("Ментор", "ID", "Курс", "Событие", "Начало", "Конец"))
        for cell in ws1[1]:
            cell.font = Font(bold=True)
        for r in rows_detail:
            ws1.append((r["mentor"], r["mentor_id"], r["course"], r["title"], r["start"], r["end"]))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fn = f"mentors-salary-{d0.strftime('%d.%m.%Y')}-{d1.strftime('%d.%m.%Y')}.xlsx"
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{fn}"'
        return resp

    def salary_create(self, request):
        if not self._can_manage_mentors(request):
            return HttpResponseForbidden()
        if request.method == "POST":
            mentor_id = (request.POST.get("mentor") or "").strip()
            course_id = (request.POST.get("course") or "").strip()
            amount = (request.POST.get("amount") or "").strip()
            comment = (request.POST.get("comment") or "").strip()
            try:
                salary = Salary(
                    mentor_id=int(mentor_id),
                    course_id=int(course_id) if course_id else None,
                    amount=Decimal(amount),
                    comment=comment,
                )
                salary.save()
                messages.success(request, "Выплата добавлена.")
            except (ValueError, InvalidOperation):
                messages.error(request, "Укажите корректную сумму.")
            except Exception as exc:
                messages.error(request, f"Ошибка: {exc}")
        return redirect(reverse(f"{self.name}:mentors_salary"))

    def salary_delete(self, request, salary_id: int):
        if not self._can_manage_mentors(request):
            return HttpResponseForbidden()
        salary = get_object_or_404(Salary, pk=salary_id)
        salary.delete()
        messages.success(request, "Выплата удалена.")
        return redirect(reverse(f"{self.name}:mentors_salary"))

    def mentor_quick_create(self, request):
        if not self._can_manage_mentors(request):
            return HttpResponseForbidden()
        if request.method != "POST":
            return redirect(reverse(f"{self.name}:settings_mentor_changelist"))
        cl_url = reverse(f"{self.name}:settings_mentor_changelist")
        first_name = (request.POST.get("first_name") or "").strip()
        last_name = (request.POST.get("last_name") or "").strip()
        email = (request.POST.get("email") or "").strip()
        phone = (request.POST.get("phone_number") or "").strip()
        if not first_name or not last_name:
            messages.error(request, "Укажите имя и фамилию.")
            return redirect(cl_url)

        def _parse_dec(val, default=None):
            raw = (val or "").strip().replace(",", ".")
            if raw == "":
                return default
            try:
                return Decimal(raw)
            except InvalidOperation:
                return None

        payment_form = (request.POST.get("payment_form") or Mentor.PaymentForm.FIXED).strip()
        if payment_form not in Mentor.PaymentForm:
            payment_form = Mentor.PaymentForm.FIXED
        fixed_raw = request.POST.get("fixed_rate")
        fixed_rate = _parse_dec(fixed_raw, Decimal("0"))
        if fixed_rate is None:
            messages.error(request, "Укажите корректную фиксированную ставку.")
            return redirect(cl_url)
        payment_rate = _parse_dec(request.POST.get("payment_rate"), None)
        percentage_rate = _parse_dec(request.POST.get("percentage_rate"), None)

        middle_name = (request.POST.get("middle_name") or "").strip()
        skills = (request.POST.get("skills") or "").strip()
        workplace = (request.POST.get("workplace") or "").strip()
        documents_folder = (request.POST.get("documents_folder") or "").strip()
        note = (request.POST.get("note") or "").strip()
        departure_reason = (request.POST.get("departure_reason") or "").strip()

        birth_date = parse_date((request.POST.get("birth_date") or "").strip())
        departure_date = parse_date((request.POST.get("departure_date") or "").strip())

        base_user = slugify(f"{first_name}-{last_name}")[:40] or "mentor"
        username = base_user
        n = 0
        while User.objects.filter(username=username).exists():
            n += 1
            username = f"{base_user}{n}"
        try:
            with transaction.atomic():
                user = User(
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    email=email or "",
                    phone_number=phone or "",
                    role="Ментор",
                    is_staff=False,
                )
                raw_password = _require_valid_password(request.POST.get("password"), user)
                user.set_password(raw_password)
                user.save()
                mentor = Mentor(
                    user=user,
                    middle_name=middle_name,
                    birth_date=birth_date,
                    skills=skills,
                    workplace=workplace,
                    documents_folder=documents_folder,
                    payment_form=payment_form,
                    payment_rate=payment_rate,
                    fixed_rate=fixed_rate,
                    percentage_rate=percentage_rate,
                    note=note,
                    departure_date=departure_date,
                    departure_reason=departure_reason,
                )
                f = request.FILES.get("contract_file")
                if f:
                    mentor.contract_file = f
                mentor.full_clean()
                mentor.save()
        except ValidationError as exc:
            messages.error(request, _validation_message(exc))
            return redirect(cl_url)
        except Exception as exc:
            messages.error(request, f"Не удалось создать ментора: {exc}")
            return redirect(cl_url)
        messages.success(request, f"Ментор создан. Логин: {username}")
        return redirect(cl_url)

    def student_quick_create(self, request):
        if request.method != "POST":
            return redirect(reverse(f"{self.name}:settings_student_changelist"))
        cl_url = reverse(f"{self.name}:settings_student_changelist")
        first_name = (request.POST.get("first_name") or "").strip()
        last_name = (request.POST.get("last_name") or "").strip()
        phone = (request.POST.get("phone_number") or "").strip()
        email = (request.POST.get("email") or "").strip()
        middle_name = (request.POST.get("middle_name") or "").strip()
        gender = (request.POST.get("gender") or "").strip()
        birth_date = parse_date((request.POST.get("birth_date") or "").strip())
        telegram_nick = (request.POST.get("telegram_nick") or "").strip()
        from_where = (request.POST.get("from_where") or "").strip()
        documents_folder = (request.POST.get("documents_folder") or "").strip()
        parent_phone = (request.POST.get("parent_phone") or "").strip()
        note = (request.POST.get("note") or "").strip()
        status = (request.POST.get("status") or "active").strip()
        course_id = (request.POST.get("course_id") or "").strip()

        if not first_name or not last_name:
            messages.error(request, "Укажите имя и фамилию.")
            return redirect(cl_url)
        base_user = slugify(f"{first_name}-{last_name}")[:40] or "student"
        username = base_user
        n = 0
        while User.objects.filter(username=username).exists():
            n += 1
            username = f"{base_user}{n}"
        try:
            with transaction.atomic():
                user = User(
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    phone_number=phone or "",
                    email=email or "",
                    role="Студент",
                    is_staff=False,
                )
                raw_password = _require_valid_password(request.POST.get("password"), user)
                user.set_password(raw_password)
                user.save()
                student = Student.objects.create(
                    user=user,
                    status=status if status in {"active", "inactive", "left", "frozen"} else "active",
                    middle_name=middle_name,
                    gender=gender,
                    birth_date=birth_date,
                    telegram_nick=telegram_nick,
                    from_where=from_where,
                    documents_folder=documents_folder,
                    parent_phone=parent_phone,
                    note=note,
                )
                if course_id and course_id.isdigit():
                    course = Cursues.objects.filter(pk=int(course_id)).first()
                    if course:
                        Enrollment.objects.get_or_create(student=student, course=course)
                        course.students.add(student)
        except ValidationError as exc:
            messages.error(request, _validation_message(exc))
            return redirect(cl_url)
        except Exception as exc:
            messages.error(request, f"Не удалось создать студента: {exc}")
            return redirect(cl_url)
        messages.success(request, f"Студент создан. Логин: {username}")
        return redirect(cl_url)

    def parent_quick_create(self, request):
        if request.method != "POST":
            # GET-запрос → показываем форму с поиском студентов
            students_qs = Student.objects.select_related("user").filter(is_archived=False)
            org_id = request.session.get("current_org_id")
            if org_id:
                students_qs = students_qs.filter(organization_id=org_id)
            students_qs = students_qs.order_by("user__last_name", "user__first_name", "user__username")
            context = {
                "title": "Создать родителя",
                "students": students_qs,
                **self.each_context(request),
            }
            return TemplateResponse(request, "admin/parent_create.html", context)
        
        cl_url = reverse(f"{self.name}:settings_parent_changelist")
        first_name = (request.POST.get("first_name") or "").strip()
        last_name = (request.POST.get("last_name") or "").strip()
        phone = (request.POST.get("phone_number") or "").strip()
        student_ids = request.POST.getlist("students")
        
        if not first_name or not last_name:
            messages.error(request, "Укажите имя и фамилию.")
            return redirect(cl_url)
        
        if not student_ids:
            messages.error(request, "Выберите хотя бы одного студента для привязки.")
            return redirect(cl_url)

        students_qs = Student.objects.filter(pk__in=student_ids, is_archived=False)
        org_id = request.session.get("current_org_id")
        if org_id:
            students_qs = students_qs.filter(organization_id=org_id)
        if not students_qs.exists():
            messages.error(
                request,
                "Выбранные студенты не найдены. Проверьте организацию в боковом меню и список студентов.",
            )
            return redirect(cl_url)

        base_user = slugify(f"{first_name}-{last_name}")[:40] or "parent"
        username = base_user
        n = 0
        while User.objects.filter(username=username).exists():
            n += 1
            username = f"{base_user}{n}"
        
        try:
            with transaction.atomic():
                user = User(
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    phone_number=phone or "",
                    role="Родитель",
                    is_staff=False,
                )
                raw_password = _require_valid_password(request.POST.get("password"), user)
                user.set_password(raw_password)
                user.save()
                parent = Parent.objects.create(
                    user=user,
                    phone_number=phone or "",
                    organization_id=org_id or None,
                )
                parent.students.set(students_qs)
        except ValidationError as exc:
            messages.error(request, _validation_message(exc))
            return redirect(cl_url)
        except Exception as exc:
            messages.error(request, f"Не удалось создать родителя: {exc}")
            return redirect(cl_url)
        
        messages.success(request, f"Родитель создан. Логин: {username}")
        return redirect(cl_url)

    def students_csv(self, request):
        qs = Student.objects.select_related("user")
        status = request.GET.get("status")
        if status:
            qs = qs.filter(status=status)
        archived = request.GET.get("archived")
        if archived == "1":
            qs = qs.filter(is_archived=True)
        elif archived == "0":
            qs = qs.filter(is_archived=False)
        q = (request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(user__username__icontains=q) | qs.filter(user__first_name__icontains=q) | qs.filter(
                user__last_name__icontains=q
            )

        lines = ["ID,ФИО,Телефон,Статус"]
        for s in qs.order_by("id")[:5000]:
            u = s.user
            full_name = (u.get_full_name() or u.username).replace(",", " ")
            phone = (u.phone_number or "").replace(",", " ")
            lines.append(f"{s.id},{full_name},{phone},{s.get_status_display()}")
        content = "\n".join(lines)
        resp = HttpResponse(content, content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = 'attachment; filename="students.csv"'
        return resp

    def students_pdf(self, request):
        qs = Student.objects.select_related("user")
        status = request.GET.get("status")
        if status:
            qs = qs.filter(status=status)
        archived = request.GET.get("archived")
        if archived == "1":
            qs = qs.filter(is_archived=True)
        elif archived == "0":
            qs = qs.filter(is_archived=False)
        q = (request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(user__username__icontains=q) | qs.filter(user__first_name__icontains=q) | qs.filter(
                user__last_name__icontains=q
            )

        rows = []
        for s in qs.order_by("id")[:1000]:
            u = s.user
            rows.append(
                [
                    str(s.id),
                    (u.get_full_name() or u.username),
                    (u.phone_number or ""),
                    s.get_status_display(),
                ]
            )

        content = _simple_pdf_table(
            title="Students",
            headers=["ID", "Name", "Phone", "Status"],
            rows=rows,
        )
        resp = HttpResponse(content, content_type="application/pdf")
        resp["Content-Disposition"] = 'attachment; filename="students.pdf"'
        return resp

    def student_drawer(self, request, student_id: int):
        student = Student.objects.select_related("user").get(pk=student_id)
        course_pk = request.GET.get("course")
        course_pk_int = int(course_pk) if course_pk and str(course_pk).isdigit() else None

        enrollment = None
        if course_pk_int:
            enrollment = (
                Enrollment.objects.select_related("course")
                .filter(student=student, course_id=course_pk_int, course__is_archived=False)
                .first()
            )
        if not enrollment:
            enrollment = (
                Enrollment.objects.select_related("course")
                .filter(student=student, course__is_archived=False)
                .order_by("-created_at")
                .first()
            )

        paid_total = 0.0
        tuition_amount = 0.0
        course = None
        contract = None
        if enrollment:
            course = enrollment.course
            tuition_amount = float(enrollment.tuition_amount or 0)
            paid_total = float(enrollment.paid_total or 0)
            contract = CourseContract.objects.filter(course=course, student=student).first()

        months_total = 0
        months_paid = 0
        if course:
            price = float(course.price or 0)
            if tuition_amount and price > 0:
                months_total = max(1, int(round(tuition_amount / price)))
                months_paid = min(months_total, int(round(paid_total / price)))
            else:
                months_total = max(1, int(round((course.duration_days or 30) / 30)))
                months_paid = min(months_total, int(round(paid_total / price))) if price > 0 else 0

        payments_qs = Payment.objects.filter(student=student).select_related("course").order_by("-created_at")
        if course:
            payments_qs = payments_qs.filter(course=course)
        payments_total = float(payments_qs.aggregate(total=Sum("amount"))["total"] or 0)
        page_size = 10
        paginator = Paginator(payments_qs, page_size)
        page_number = request.GET.get("p") or 1
        payments_page = paginator.get_page(page_number)

        first_payment_qs = payments_qs.order_by("created_at")
        first_payment = first_payment_qs.first() if course else None
        first_visit_display = "Неизвестно"
        if first_payment and first_payment.created_at:
            first_visit_display = timezone.localdate(first_payment.created_at).strftime("%d.%m.%Y")

        added_by_name = "—"
        if course:
            m = course.mentors.select_related("user").first()
            if m and m.user:
                added_by_name = m.user.get_full_name() or m.user.username

        month_slots = [{"paid": i < months_paid} for i in range(months_total)] if months_total else []
        progress_pct = int(round(100 * months_paid / months_total)) if months_total else 0

        phone = (student.user.phone_number or "").strip()
        wa_digits = re.sub(r"\D+", "", phone)
        whatsapp_href = f"https://wa.me/{wa_digits}" if wa_digits else ""

        instant_pay_url = ""
        if course:
            host_base = request.build_absolute_uri("/").rstrip("/")
            instant_pay_url = f"{host_base}/instant-pay/?student={student.pk}&course={course.pk}"

        context = {
            "student": student,
            "user": student.user,
            "enrollment": enrollment,
            "course": course,
            "contract": contract,
            "tuition_amount": tuition_amount,
            "paid_total": paid_total,
            "debt_amount": max(0.0, tuition_amount - paid_total),
            "months_total": months_total,
            "months_paid": months_paid,
            "month_slots": month_slots,
            "progress_pct": progress_pct,
            "payments_page": payments_page,
            "payments_total": payments_total,
            "first_visit_display": first_visit_display,
            "added_by_name": added_by_name,
            "whatsapp_href": whatsapp_href,
            "instant_pay_url": instant_pay_url,
            "schedule_note": (course.schedule_note if course else "") or "—",
            "course_start_display": course.start.strftime("%d.%m.%Y") if course and course.start else "—",
            **self.each_context(request),
        }
        return TemplateResponse(request, "admin/student_drawer.html", context)

    def index(self, request, extra_context=None):
        context = extra_context or {}

        today = timezone.localdate()
        start_month = today.replace(day=1)
        start_year = today.replace(month=1, day=1)
        start_week = today - timedelta(days=6)
        first_prev_month = (start_month - timedelta(days=1)).replace(day=1)
        prev_month_end = start_month - timedelta(days=1)

        leads_qs = Lead.objects.all()
        students_qs = Student.objects.filter(is_archived=False)
        tuition_qs = TuitionPayment.objects.all()
        payments_qs = Payment.objects.all()
        salary_qs = Salary.objects.all()
        drops_qs = CourseDrop.objects.all()

        license_alert = None
        bill = (
            BillingRecord.objects.filter(status=BillingRecord.Status.ACTIVE, expires_at__isnull=False)
            .order_by("expires_at")
            .first()
        )
        if bill and bill.expires_at:
            days_left = (bill.expires_at - today).days
            if days_left <= 30:
                if days_left <= 0:
                    msg = "Лицензия на использование системы истекла"
                elif days_left == 1:
                    msg = "Лицензия на использование системы истечет через 1 день"
                else:
                    msg = f"Лицензия на использование системы истечет через {days_left} дней"
                license_alert = {"show": True, "message": msg, "days_left": days_left}

        this_month_tuition = float(
            tuition_qs.filter(created_at__date__gte=start_month, created_at__date__lte=today).aggregate(
                t=Sum("amount")
            )["t"]
            or 0
        )
        prev_month_tuition = float(
            tuition_qs.filter(
                created_at__date__gte=first_prev_month, created_at__date__lte=prev_month_end
            ).aggregate(t=Sum("amount"))["t"]
            or 0
        )
        if prev_month_tuition > 0:
            pct_vs_prev = round(100.0 * (this_month_tuition - prev_month_tuition) / prev_month_tuition, 1)
            ring_pct = min(100, int(round(100.0 * this_month_tuition / prev_month_tuition)))
        else:
            pct_vs_prev = 100.0 if this_month_tuition > 0 else 0.0
            ring_pct = 100 if this_month_tuition > 0 else 0

        money_paid_today = float(payments_qs.filter(created_at__date=today).aggregate(total=Sum("amount"))["total"] or 0)
        money_paid_week = float(
            payments_qs.filter(created_at__date__gte=start_week).aggregate(total=Sum("amount"))["total"] or 0
        )
        money_paid_month = float(
            payments_qs.filter(created_at__date__gte=start_month).aggregate(total=Sum("amount"))["total"] or 0
        )
        money_paid_year = float(
            payments_qs.filter(created_at__date__gte=start_year).aggregate(total=Sum("amount"))["total"] or 0
        )

        leads_today = leads_qs.filter(created_at__date=today).count()
        leads_week = leads_qs.filter(created_at__date__gte=start_week).count()
        leads_month = leads_qs.filter(created_at__date__gte=start_month).count()
        leads_year = leads_qs.filter(created_at__date__gte=start_year).count()

        students_today = students_qs.filter(created_at__date=today).count()
        students_week = students_qs.filter(created_at__date__gte=start_week).count()
        students_month = students_qs.filter(created_at__date__gte=start_month).count()
        students_year = students_qs.filter(created_at__date__gte=start_year).count()

        drops_today = drops_qs.filter(dropped_at=today).count()
        drops_week = drops_qs.filter(dropped_at__gte=start_week).count()
        drops_month = drops_qs.filter(dropped_at__gte=start_month).count()
        drops_year = drops_qs.filter(dropped_at__gte=start_year).count()

        courses_income = (
            payments_qs.values("course_id", "course__title", "course__subject")
            .annotate(income=Sum("amount"))
            .order_by("-income")
        )
        courses_income = [c for c in courses_income if c.get("course_id")][:8]
        course_ids = [int(c["course_id"]) for c in courses_income]
        courses_by_id = {c.id: c for c in Cursues.objects.filter(pk__in=course_ids).prefetch_related("mentors")}
        courses_top = []
        for row in courses_income:
            cid = int(row["course_id"])
            course = courses_by_id.get(cid)
            income = float(row["income"] or 0)
            mentor_ids = list(course.mentors.values_list("id", flat=True)) if course else []
            salary_total = float(
                Salary.objects.filter(mentor_id__in=mentor_ids).aggregate(s=Sum("amount"))["s"] or 0
            )
            profit = income - salary_total
            courses_top.append(
                {
                    "title": row.get("course__title") or "—",
                    "subtitle": (row.get("course__subject") or "").strip(),
                    "income": income,
                    "salary": salary_total,
                    "profit": profit,
                    "income_compact": _compact_money(income),
                    "salary_compact": _compact_money(salary_total),
                    "profit_compact": _compact_money(profit),
                    "income_display": _format_money_ru(income),
                    "salary_display": _format_money_ru(salary_total),
                    "profit_display": _format_money_ru(profit),
                }
            )

        income_12m_s = _sum_by_month(payments_qs, months=12, field="amount")
        expense_12m_s = _sum_by_month(salary_qs, months=12, field="amount")
        payments_6m_s = _sum_by_month(payments_qs, months=6, field="amount")
        leads_6m_s = _count_by_month(leads_qs, months=6)
        students_6m_s = _distinct_paying_students_by_month(payments_qs, 6)
        visits_30d_s = _calendar_events_by_day(30)

        sum_in_12 = float(sum(income_12m_s.values))
        sum_ex_12 = float(sum(expense_12m_s.values))
        sum_pay_6 = float(sum(payments_6m_s.values))
        sum_leads_6 = int(round(sum(leads_6m_s.values)))
        sum_stu_6 = int(round(sum(students_6m_s.values)))
        sum_vis_30 = int(round(sum(visits_30d_s.values)))

        payments_all_time = float(payments_qs.aggregate(t=Sum("amount"))["t"] or 0)
        salaries_all_time = float(salary_qs.aggregate(t=Sum("amount"))["t"] or 0)

        series = {
            "income": income_12m_s.__dict__,
            "expense": expense_12m_s.__dict__,
            "payments_6m": payments_6m_s.__dict__,
            "leads_6m": leads_6m_s.__dict__,
            "students_6m": students_6m_s.__dict__,
            "visits_30d": visits_30d_s.__dict__,
        }

        def _fmt_card(d: dict) -> dict:
            return {k: _format_int_ru(v) for k, v in d.items()}

        context.update(
            {
                "license_alert": license_alert,
                "overview": {
                    "payments_all": _format_money_ru(payments_all_time),
                    "salaries_all": _format_money_ru(salaries_all_time),
                    "net_all": _format_money_ru(payments_all_time - salaries_all_time),
                    "students": _format_int_ru(students_qs.count()),
                    "leads": _format_int_ru(leads_qs.count()),
                },
                "income_block": {
                    "this_month": this_month_tuition,
                    "this_month_compact": _compact_money(this_month_tuition),
                    "this_month_display": _format_money_ru(this_month_tuition),
                    "prev_month": prev_month_tuition,
                    "prev_month_display": _format_money_ru(prev_month_tuition),
                    "pct_vs_prev": pct_vs_prev,
                    "ring_pct": ring_pct,
                },
                "kpi": {
                    "students": students_qs.count(),
                    "mentors": Mentor.objects.count(),
                    "leads": leads_qs.count(),
                    "courses": Cursues.objects.filter(is_archived=False).count(),
                    "users": User.objects.count(),
                    "staff": User.objects.filter(is_staff=True).count(),
                },
                "kpi_display": {
                    "students": _format_int_ru(students_qs.count()),
                    "mentors": _format_int_ru(Mentor.objects.count()),
                    "leads": _format_int_ru(leads_qs.count()),
                    "courses": _format_int_ru(Cursues.objects.filter(is_archived=False).count()),
                    "users": _format_int_ru(User.objects.count()),
                    "staff": _format_int_ru(User.objects.filter(is_staff=True).count()),
                },
                "courses_top": courses_top,
                "cards": {
                    "leads": {"today": leads_today, "week": leads_week, "month": leads_month, "year": leads_year},
                    "students": {
                        "today": students_today,
                        "week": students_week,
                        "month": students_month,
                        "year": students_year,
                    },
                    "drops": {"today": drops_today, "week": drops_week, "month": drops_month, "year": drops_year},
                },
                "cards_display": {
                    "leads": _fmt_card(
                        {"today": leads_today, "week": leads_week, "month": leads_month, "year": leads_year}
                    ),
                    "students": _fmt_card(
                        {
                            "today": students_today,
                            "week": students_week,
                            "month": students_month,
                            "year": students_year,
                        }
                    ),
                    "drops": _fmt_card(
                        {"today": drops_today, "week": drops_week, "month": drops_month, "year": drops_year}
                    ),
                },
                "money": {
                    "paid": {
                        "today": money_paid_today,
                        "week": money_paid_week,
                        "month": money_paid_month,
                        "year": money_paid_year,
                    }
                },
                "money_display": {
                    "paid": {
                        "today": _format_money_ru(money_paid_today),
                        "week": _format_money_ru(money_paid_week),
                        "month": _format_money_ru(money_paid_month),
                        "year": _format_money_ru(money_paid_year),
                    }
                },
                "chart_totals": {
                    "students_6m": _format_int_ru(sum_stu_6),
                    "visits_30d": _format_int_ru(sum_vis_30),
                    "payments_6m": _format_money_ru(sum_pay_6),
                    "leads_6m": _format_int_ru(sum_leads_6),
                    "income_12m": _format_money_ru(sum_in_12),
                    "expense_12m": _format_money_ru(sum_ex_12),
                    "net_12m": _format_money_ru(sum_in_12 - sum_ex_12),
                },
                "series": series,
                "dashboard_month_title": formats.date_format(today, "F Y"),
            }
        )

        return super().index(request, extra_context=context)


def _pdf_escape(text: str) -> str:
    return (
        text.replace("\\", "\\\\")
        .replace("(", "\\(")
        .replace(")", "\\)")
        .replace("\r", " ")
        .replace("\n", " ")
    )


def _simple_pdf_table(title: str, headers: list[str], rows: list[list[str]]) -> bytes:
    # Minimal PDF generator (single font: Helvetica). Good enough for demo exports.
    # Page: A4 portrait, 595x842 points.
    width, height = 595, 842
    margin_x, margin_y = 40, 48
    line_h = 14
    y = height - margin_y

    lines: list[str] = []
    lines.append(f"/F1 16 Tf {margin_x} {y} Td ({_pdf_escape(title)}) Tj")
    y -= 24

    def emit_row(cells: list[str], bold: bool = False):
        nonlocal y
        if y < margin_y + 60:
            # new page marker
            lines.append("__NEW_PAGE__")
            y = height - margin_y
        font_size = 10
        lines.append(f"/F1 {font_size} Tf {margin_x} {y} Td ({_pdf_escape(' | '.join(cells))}) Tj")
        y -= line_h

    emit_row(headers)
    emit_row(["-" * 80])
    for r in rows:
        emit_row(r)

    # Split into pages
    pages: list[list[str]] = [[]]
    for l in lines:
        if l == "__NEW_PAGE__":
            pages.append([])
        else:
            pages[-1].append(l)

    objects: list[bytes] = []

    def add_obj(data: str) -> int:
        objects.append(data.encode("utf-8"))
        return len(objects)

    # Catalog + Pages + Font
    font_obj = add_obj("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")

    page_objs: list[int] = []
    content_objs: list[int] = []

    pages_kids = []
    for page_lines in pages:
        stream = "BT\n" + "\n".join(page_lines) + "\nET\n"
        content_obj = add_obj(f"<< /Length {len(stream.encode('utf-8'))} >>\nstream\n{stream}endstream")
        content_objs.append(content_obj)
        page_obj = add_obj(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {width} {height}] "
            f"/Resources << /Font << /F1 {font_obj} 0 R >> >> /Contents {content_obj} 0 R >>"
        )
        page_objs.append(page_obj)
        pages_kids.append(f"{page_obj} 0 R")

    pages_obj = add_obj(f"<< /Type /Pages /Kids [ {' '.join(pages_kids)} ] /Count {len(page_objs)} >>")
    # ensure Pages is object 2 (we referenced it); if not, patch by inserting dummy
    # Simpler: rebuild with fixed ordering if needed.
    if pages_obj != 2:
        # rebuild objects with fixed order: [Catalog, Pages, Font, contents..., pages...]
        objects = []
        # placeholders for indices
        catalog_idx = add_obj("<< /Type /Catalog /Pages 2 0 R >>")
        pages_idx = add_obj("<< >>")  # temp
        font_idx = add_obj("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
        page_objs = []
        pages_kids = []
        for page_lines in pages:
            stream = "BT\n" + "\n".join(page_lines) + "\nET\n"
            content_obj = add_obj(f"<< /Length {len(stream.encode('utf-8'))} >>\nstream\n{stream}endstream")
            page_obj = add_obj(
                f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {width} {height}] "
                f"/Resources << /Font << /F1 {font_idx} 0 R >> >> /Contents {content_obj} 0 R >>"
            )
            page_objs.append(page_obj)
            pages_kids.append(f"{page_obj} 0 R")
        objects[pages_idx - 1] = f"<< /Type /Pages /Kids [ {' '.join(pages_kids)} ] /Count {len(page_objs)} >>".encode(
            "utf-8"
        )
    else:
        # add catalog after pages to keep references fine
        add_obj("<< /Type /Catalog /Pages 2 0 R >>")

    # build xref
    out = bytearray()
    out.extend(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for i, obj in enumerate(objects, start=1):
        offsets.append(len(out))
        out.extend(f"{i} 0 obj\n".encode("utf-8"))
        out.extend(obj)
        out.extend(b"\nendobj\n")

    xref_start = len(out)
    out.extend(f"xref\n0 {len(objects)+1}\n".encode("utf-8"))
    out.extend(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        out.extend(f"{off:010d} 00000 n \n".encode("utf-8"))
    out.extend(
        f"trailer\n<< /Size {len(objects)+1} /Root {len(objects)} 0 R >>\nstartxref\n{xref_start}\n%%EOF\n".encode(
            "utf-8"
        )
    )
    return bytes(out)


crm_admin_site = CRMAdminSite(name="crm_admin")
